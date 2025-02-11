import threading
import time
from typing import Type, Callable
import os

import pandas as pd
from PySide6 import QtWidgets, QtCore, QtGui
from openpyxl import load_workbook

from creator.controllers.config_controller import ConfigController
from creator.controllers.db_controller import DBController
from creator.controllers.api_controller import APIController, APIException
from creator.utils.sip_status import SIPStatus
from creator.utils.state import State
from creator.utils.path_loader import resource_path
from creator.widgets.warning_dialog import WarningDialog


class SIPStatusThread(threading.Thread):
    def __init__(self, state: State):
        super().__init__()

        self.state = state

        self.daemon = True
        self.start()

    def run(self):
        # On first launch, allow the application time to start up
        time.sleep(30)

        while True:
            try:
                self._check_sip_status()
            except APIException:
                pass

            time.sleep(60)

    def _check_sip_status(self):
        # TODO: better sql get for only uploaded sips
        for sip in self.state.sips:
            # We only care about unfinished edepot statusses
            if sip.status in (
                SIPStatus.ACCEPTED,
                SIPStatus.REJECTED,
                SIPStatus.IN_PROGRESS,
                SIPStatus.SIP_CREATED,
                SIPStatus.UPLOADING,
                SIPStatus.DELETED,
            ):
                continue

            # Make sure we have the edepot id
            if sip.edepot_sip_id is None:
                sip.set_edepot_sip_id(APIController.get_sip_id(sip))

            new_status, fail_reason = APIController.get_sip_status(sip)
            # new_status = APIController.get_sip_status_from_dossiers(sip)

            if new_status == sip.status:
                continue

            sip.set_status(new_status)

            self.state.update_sip(sip)


class Application(QtWidgets.QApplication, QtCore.QObject):
    type_changed: QtCore.Signal = QtCore.Signal()

    def __init__(self, mainwindow: Type[QtWidgets.QMainWindow], set_main_callback: Callable):
        super().__init__()

        self.setWindowIcon(QtGui.QIcon(resource_path("logo.ico")))

        self.db_controller = DBController("sqlite.db")
        self.config_controller = ConfigController("configuration.json")

        self.state = State(
            configuration_callback=self.config_controller.get_configuration,
            db_controller=self.db_controller,
        )

        self.db_controller.set_application()

        self.state.configuration.create_locations()

        self.sip_status_thread = SIPStatusThread(self.state)

        self.ui = mainwindow()
        self.ui.resize(800, 600)

        set_main_callback(self, self.ui)

        self.type_changed.connect(lambda: set_main_callback(self, self.ui))

        controle_list_path = self.state.configuration.misc.bestandscontrole_lijst_location
        self.bestands_controle_lijst_controller = BestandsControleLijstController(controle_list_path=controle_list_path)

    def start(self) -> None:
        self.ui.show()

    def reset_bestandscontrole_location(self) -> None:
        controle_list_path = self.state.configuration.misc.bestandscontrole_lijst_location

        if controle_list_path == self.bestands_controle_lijst_controller.controle_list_path:
            return

        self.bestands_controle_lijst_controller = BestandsControleLijstController(controle_list_path=controle_list_path)


class BestandsControleLijstController:
    def __init__(self, controle_list_path: str):
        self.controle_list_path = controle_list_path

        self.list_sheet_name = "Lijst"

        self.list_name_column = "Lijst benaming"
        self.list_start_column = "Locatie dozen: begin\n(verdiep/blok*rij*rek*ligger)"
        self.list_end_column = "Locatie dozen: eind\n(verdiep/blok*rij*rek*ligger)"
        self.list_type_column = "Doostype\n(1, 2, 3)\n(indien meerdere types: meest voorkomende)"

        self.list_column_names = [
            self.list_name_column,
            self.list_start_column,
            self.list_end_column,
            self.list_type_column,
        ]


        self.doos_sheet_name = "Doostypes"

        self.doos_number_column = "Doosnummer"
        self.doos_type_column = "Naam doostype migratie"

        self.doos_column_names = [
            self.doos_number_column,
            self.doos_type_column,
        ]

        self.valid = False
        self.list_df = self._load_df()

    def _check_valid_controle_list(self) -> bool:
        if self.controle_list_path in ("", "."):
            # NOTE: probably not set in config, don't show a message yet
            return False

        # 0. Check if file exists
        if not os.path.exists(self.controle_list_path):
            WarningDialog(
                title="Bestandscontrole niet gevonden",
                text=f"Bestandscontrole niet gevonden op locatie '{self.controle_list_path}'."
            ).exec()
            return False
        
        # 1. Check if sheet exists
        wb = load_workbook(
            self.controle_list_path,
            read_only=True,
            data_only=True,
            keep_links=False,
            rich_text=False,
        )

        if not self.list_sheet_name in wb.sheetnames:
            WarningDialog(
                title="Tab niet gevonden in bestandscontrole",
                text=f"Tab '{self.list_sheet_name}' niet gevonden in bestandscontrole."
            ).exec()
            wb.close()
            return False
        
        if not self.doos_sheet_name in wb.sheetnames:
            WarningDialog(
                title="Tab niet gevonden in bestandscontrole",
                text=f"Tab '{self.doos_sheet_name}' niet gevonden in bestandscontrole."
            ).exec()
            wb.close()
            return False

        ws_list = wb[self.list_sheet_name]
        ws_doos = wb[self.doos_sheet_name]
        list_data = ws_list.values
        doos_data = ws_doos.values

        # 2. Check if headers exist in list data
        try:
            # TODO: check all required columns?
            # NOTE: we know it is row number 3, but still, check it this way
            while "Datum" not in (headers := next(list_data)):
                pass
        except StopIteration:
            WarningDialog(
                title="Verwachte hoofdingen niet gevonden",
                text=f"De verwachte hoofdingen waren niet gevonden in de bestandscontrole in tab '{self.list_sheet_name}'."
            ).exec()
            wb.close()
            return False
        
        # 3. Check if all headers present in list data
        headers = [h for h in headers if h is not None]

        missing_columns = []

        for h in self.list_column_names:
            if h not in headers:
                missing_columns.append(h)

        if len(missing_columns) > 0:
            WarningDialog(
                title="Verwachte hoofdingen niet gevonden",
                text=f"De verwachte hoofdingen waren niet gevonden in de bestandscontrole in tab '{self.list_sheet_name}'."
            ).exec()
            wb.close()
            return False

        # 4. Check if headers exist in doos data
        try:
            # TODO: check all required columns?
            # NOTE: we know it is row number 6, but still, check it this way
            while "Doosnummer" not in (headers := next(doos_data)):
                pass
        except StopIteration:
            WarningDialog(
                title="Verwachte hoofdingen niet gevonden",
                text=f"De verwachte hoofdingen waren niet gevonden in de bestandscontrole in tab '{self.doos_sheet_name}'."
            ).exec()
            wb.close()
            return False
        
        # 5. Check if all headers present in doos data
        headers = [h for h in headers if h is not None]

        missing_columns = []

        for h in self.doos_column_names:
            if h not in headers:
                missing_columns.append(h)

        if len(missing_columns) > 0:
            WarningDialog(
                title="Verwachte hoofdingen niet gevonden",
                text=f"De verwachte hoofdingen waren niet gevonden in de bestandscontrole in tab '{self.doos_sheet_name}'."
            ).exec()
            wb.close()
            return False
        
        wb.close()
        return True

    def _load_df(self) -> pd.DataFrame:
        if not self._check_valid_controle_list():
            # NOTE: if the list is not valid, we cannot continue
            self.valid = False
            return
        self.valid = True

        wb = load_workbook(
            self.controle_list_path,
            read_only=True,
            data_only=True,
            keep_links=False,
            rich_text=False,
        )

        ws_list = wb[self.list_sheet_name]
        list_data = list(ws_list.values)

        # NOTE: only take the parts we need
        list_data_header_row = [i for i, r in enumerate(list_data) if all(c in r for c in self.list_column_names)][0]
        list_data_columns = [i for i, c in enumerate(list_data[list_data_header_row]) if c in self.list_column_names]
        list_data = [[c for i, c in enumerate(r) if i in list_data_columns] for r in list_data[list_data_header_row+1:]]

        list_df = pd.DataFrame(
            [r for r in list_data if not all(c is None for c in r)],
            columns=self.list_column_names,
        ).fillna("").astype(str).convert_dtypes()
        del list_data
        
        # NOTE: same for doos
        ws_doos = wb[self.doos_sheet_name]
        doos_data = list(ws_doos.values)

        doos_data_header_row = [i for i, r in enumerate(doos_data) if all(c in r for c in self.doos_column_names)][0]
        doos_data_columns = [i for i, c in enumerate(doos_data[doos_data_header_row]) if c in self.doos_column_names]
        doos_data = [[c for i, c in enumerate(r) if i in doos_data_columns] for r in doos_data[doos_data_header_row+1:]]

        doos_df = pd.DataFrame(
            [r for r in doos_data if not all(c is None for c in r)],
            columns=self.doos_column_names,
        ).fillna("").astype(str).convert_dtypes()
        del doos_data

        merged_df = list_df.merge(doos_df, left_on=self.list_type_column, right_on=self.doos_number_column, how='left')
        merged_df[self.doos_type_column] = merged_df[self.doos_type_column].fillna(merged_df[self.list_type_column])

        out_df = merged_df[[self.list_name_column, self.list_start_column, self.list_end_column, self.doos_type_column]]

        wb.close()
        return out_df

    def get_values(self, overdrachtslijst_name: str) -> dict|None:
        # Try to load it again, just in case the file changed
        self._load_df()

        # NOTE: if overdrachtslijst_name contains something like _klant near the end, we want to trim all that off
        trimmed_overdrachtslijst_name = overdrachtslijst_name.split("_klant")[0]

        # Return values required as df
        output = self.list_df.loc[self.list_df[self.list_name_column] == trimmed_overdrachtslijst_name]

        if len(output) == 0:
            WarningDialog(
                title="Overdrachtslijst niet gevonden in bestandscontrole",
                text=f"Overdrachtslijst '{trimmed_overdrachtslijst_name}' is niet gevonden in de kolom '{self.list_name_column}' van de bestandscontrole."
            ).exec()
            return
        
        if len(output) != 1:
            WarningDialog(
                title="Overdrachtslijst te vaak gevonden in bestandscontrole",
                text=f"Overdrachtslijst '{trimmed_overdrachtslijst_name}' is te vaak gevonden in de kolom '{self.list_name_column}' van de bestandscontrole.\n\nEr zijn {len(output)} rijen die over dezelfde overdrachtslijst gaan."
            ).exec()
            return

        # NOTE: check if the values are present
        for col in (self.list_start_column, self.list_end_column, self.doos_type_column):
            if output.at[output.index[0], col] in (None, ""):
                WarningDialog(
                    title="Lege waarde gevonden",
                    text=f"De kolom '{col}' bevat een lege waarde voor overdrachtslijst '{trimmed_overdrachtslijst_name}'."
                ).exec()

        return output.fillna("").to_dict(orient="records")[0]

