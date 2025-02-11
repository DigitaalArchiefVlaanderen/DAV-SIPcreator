import os
import hashlib

import pandas as pd
from openpyxl import load_workbook
import shutil
import zipfile
import re

from .api_controller import APIController
from .db_controller import SIPDBController
from ..utils.state_utils.sip import SIP
from ..utils.configuration import Configuration


class FileController:
    GRID_STORAGE = "Grid"
    SIP_STORAGE = "SIPs"
    IMPORT_TEMPLATE_STORAGE = "import_templates"

    @staticmethod
    def ensure_folder_exists(location):
        if not os.path.exists(location):
            os.makedirs(location)

    @staticmethod
    def fill_import_template(df: pd.DataFrame, path: str):
        def _col_index_to_xslx_col(col_index: int) -> str:
            # NOTE: this only supports up to ZZ for now
            first_letter_value = (col_index // 26) - 1

            if first_letter_value >= 26:
                raise ValueError("There are too many columns")
            if first_letter_value == -1:
                return chr(65 + col_index)

            first_letter = chr(65 + first_letter_value)
            second_letter = chr(65 + col_index % 26)
            
            return f"{first_letter}{second_letter}"

        wb = load_workbook(path)
        ws = wb["Details"]

        for i, col in enumerate(df.columns):
            # NOTE: duplicate columns generated generate with whitespaces at the end
            # duplicate columns read generate like <col>.<n>
            re_match = re.match(r"(.*)(\.\d+| +)$", col)

            if re_match:
                col = re_match.group(1)

            ws[f"{_col_index_to_xslx_col(i)}1"] = col

        # NOTE: there is probably a better way to do this, did not find it
        for row_index, row_info in df.iterrows():
            for col_index, value in enumerate(row_info.values):
                ws[f"{_col_index_to_xslx_col(col_index)}{row_index+2}"] = value

        wb.save(path)
        wb.close()

    @staticmethod
    def _check_sip_folder_structure(sip_folder_structure: dict, df: pd.DataFrame) -> None:
        # NOTE: given the sip_folder structure, check if it still matches the df
        paths_in_sip = []

        for row in sip_folder_structure.values():
            path_in_sip, path, _type = row["Path in SIP"], row["path"], row["Type"]

            if _type == "geen":
                continue

            if sum(df["Path in SIP"] == path_in_sip) == 0:
                raise ValueError(f"Verandering in folderstructuur gevonden.\n\n'{path}' was niet aanwezig bij het aanmaken van de SIP.")
            elif sum(df["Path in SIP"] == path_in_sip) > 1:
                raise ValueError(f"Duplicate path_in_sip gevonden: '{path_in_sip}'")

            paths_in_sip.append(path_in_sip)

        for _, row in df.iterrows():
            if row["Type"] == "geen":
                continue

            if row["Path in SIP"] not in paths_in_sip:
                raise ValueError(f"Verandering in folderstructuur gevonden.\n\n'{row["Path in SIP"]}' was aanwezig bij het aanmaken van de SIP, maar is niet meer aanwezig.")

    @staticmethod
    def create_sip(configuration: Configuration, sip: SIP, df: pd.DataFrame, unfiltered_df: pd.DataFrame) -> None:
        storage_location = configuration.sips_location
        import_template_location = os.path.join(
            configuration.import_templates_location,
            f"{sip.series._id}.xlsx"
        )

        # NOTE: these locations should still exist, but just to be safe, we check
        os.makedirs(storage_location, exist_ok=True)

        if not os.path.exists(import_template_location):
            # NOTE: download the import_template again
            import_template_location = APIController.get_import_template(
                configuration=configuration,
                series_id=sip.series._id,
                environment=sip.environment,
            )

        sip_folder_structure = sip.get_sip_folder_structure()
        FileController._check_sip_folder_structure(
            sip_folder_structure=sip_folder_structure,
            df=unfiltered_df
        )

        # NOTE: make a temporary copy to save the data into
        temp_excel_location = os.path.join(
            configuration.import_templates_location,
            "temp.xlsx"
        )
        shutil.copy(
            src=import_template_location,
            dst=temp_excel_location
        )

        FileController.fill_import_template(
            df, temp_excel_location
        )

        sip_location = os.path.join(storage_location, sip.file_name)
        sidecar_location = os.path.join(storage_location, sip.sidecar_file_name)

        with zipfile.ZipFile(
            sip_location, "w", compression=zipfile.ZIP_DEFLATED
        ) as zfile:
            zfile.write(
                temp_excel_location,
                "Metadata.xlsx",
            )

            for location in sip_folder_structure.values():
                device_location = location["path"]
                path_in_sip = location["Path in SIP"]

                # Ignore bad types
                if location["Type"] != "geen":
                    zfile.write(device_location, path_in_sip)

        md5 = hashlib.md5(open(sip_location, "rb").read()).hexdigest()

        side_car_info = """
<?xml version="1.0" encoding="UTF-8"?>
<mhs:Sidecar xmlns:mhs="https://zeticon.mediahaven.com/metadata/20.3/mhs/" version="20.3" xmlns:mh="https://zeticon.mediahaven.com/metadata/20.3/mh/">
     <mhs:Technical>
              <mh:Md5>{md5}</mh:Md5>
     </mhs:Technical>
</mhs:Sidecar>""".format(
            md5=md5
        )

        with open(sidecar_location, "w", encoding="utf-8") as f:
            f.write(side_car_info)

    @staticmethod
    def existing_sip_db(configuration: Configuration, db_name: str) -> pd.DataFrame:
        path = os.path.join(configuration.sip_db_location, db_name)

        sip_db_controller = SIPDBController(path)

        if sip_db_controller.is_valid_db():
            return sip_db_controller.read_data_table()
