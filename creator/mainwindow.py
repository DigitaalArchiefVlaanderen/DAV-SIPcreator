import os
import json
import re
import threading

import zipfile
import hashlib
import ftplib
import socket
from openpyxl import load_workbook
from PySide6 import QtWidgets, QtCore, QtGui
import pandas as pd
import sqlite3 as sql
from pathlib import Path

from .application import Application

from .widgets.searchable_list_widget import (
    SearchableSelectionListView,
    SearchableListWidget,
    SIPListWidget,
)
from .widgets.tableview_widget import TableView
from .widgets.dossier_widget import DossierWidget
from .widgets.sip_widget import SIPWidget
from .widgets.toolbar import Toolbar
from .widgets.dialog import YesNoDialog, Dialog
from .widgets.warning_dialog import WarningDialog

from .controllers.file_controller import FileController
from .controllers.api_controller import APIController

from .utils.state import State
from .utils.state_utils.dossier import Dossier
from .utils.state_utils.sip import SIP
from .utils.sip_status import SIPStatus
from .utils.sqlitemodel import SQLliteModel, Color

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()

        self.application: Application = QtWidgets.QApplication.instance()
        self.state: State = self.application.state

        self.central_widget = None

        # Toolbar
        self.toolbar = Toolbar()
        self.addToolBar(self.toolbar)

    def closeEvent(self, event):
        # If the main window dies, kill the whole application
        if any(
            s.status == SIPStatus.UPLOADING
            for s in self.application.db_controller.read_sips()
        ):
            WarningDialog(
                title="Upload bezig",
                text="Waarschuwing, een upload is momenteel bezig, de applicatie kan niet gesloten worden.",
            ).exec()

            event.ignore()
            return

        event.accept()
        self.application.quit()

class DigitalWidget(QtWidgets.QWidget):
    def __init__(self, parent: MainWindow):
        super().__init__(parent)

        self.application: Application = QtWidgets.QApplication.instance()
        self.state: State = self.application.state

        self.state.sip_edepot_failed.connect(self.fail_reason_show)

    def fail_reason_show(self, sip: SIP, reason: str):
        WarningDialog(
            title="SIP upload gefaald",
            text=f"SIP '{sip.name}' is geweigerd door het E-depot met volgende reden:\n\n{reason}",
        ).exec()

        storage_location = self.state.configuration.misc.save_location
        with open(
            os.path.join(
                storage_location, FileController.SIP_STORAGE, sip.error_file_name
            ),
            "w",
            encoding="utf-8",
        ) as f:
            f.write(
                f"SIP '{sip.name}' is geweigerd door het E-depot met volgende reden:\n\n{reason}"
            )

    def setup_ui(self):
        grid_layout = QtWidgets.QGridLayout()
        self.setLayout(grid_layout)

        # Dossiers
        add_dossier_button = QtWidgets.QPushButton(text="Voeg een dossier toe")
        add_dossier_button.clicked.connect(self.add_dossier_clicked)

        add_dossiers_button = QtWidgets.QPushButton(text="Voeg folder met dossiers toe")
        add_dossiers_button.clicked.connect(
            lambda: self.add_dossier_clicked(multi=True)
        )

        self.dossiers_list_view = SearchableSelectionListView()

        grid_layout.addWidget(add_dossier_button, 0, 0)
        grid_layout.addWidget(add_dossiers_button, 0, 1)
        grid_layout.addWidget(self.dossiers_list_view, 1, 0, 1, 2)

        # SIPS
        self.create_sip_button = QtWidgets.QPushButton(text="Start SIP")
        self.create_sip_button.clicked.connect(self.create_sip_clicked)
        self.create_sip_button.setEnabled(False)
        self.sip_list_view = SIPListWidget()
        
        self.parent().toolbar.configuration_changed.connect(self.sip_list_view.reload_widgets)

        grid_layout.addWidget(self.create_sip_button, 0, 2, 1, 2)
        grid_layout.addWidget(self.sip_list_view, 1, 2, 1, 2)

    def load_items(self):
        removed_dossiers = []
        dossier_widgets = []

        for dossier in self.application.state.dossiers:
            if dossier.disabled:
                continue

            if not os.path.exists(dossier.path):
                removed_dossiers.append(dossier)
                continue

            dossier_widget = DossierWidget(dossier=dossier)

            dossier_widgets.append(dossier_widget)

        self.dossiers_list_view.add_items(
            widgets=dossier_widgets,
            selection_changed_callback=self.dossier_selection_changed,
            first_launch=True,
        )

        if len(removed_dossiers) > 0:
            dialog = YesNoDialog(
                title="Verwijderde dossiers",
                text="Een aantal dossiers lijken niet meer op hun plaats te staan.\nWilt u deze ook uit de lijst verwijderen?\n\nDeze boodschap zal anders blijven verschijnen.",
            )
            dialog.exec()

            if dialog.result():
                for dossier in removed_dossiers:
                    dossier.disabled = True
                    self.application.state.remove_dossier(dossier)

        missing_sips = []
        sips = self.application.state.sips
        sorted_sips = sorted(sips, key=lambda s: s.status.get_priority(), reverse=True)

        for sip in sorted_sips:
            # Check for missing sips
            if sip.status in (
                SIPStatus.SIP_CREATED,
                SIPStatus.UPLOADING,
                SIPStatus.UPLOADED,
                SIPStatus.ACCEPTED,
                SIPStatus.REJECTED,
            ):
                base_sip_path = os.path.join(
                    self.state.configuration.misc.save_location,
                    FileController.SIP_STORAGE,
                )
                # Check if the saved SIP and sidecar still exists
                if not os.path.exists(
                    os.path.join(
                        base_sip_path,
                        sip.file_name,
                    )
                    or not os.path.exists(
                        os.path.join(base_sip_path, sip.sidecare_file_name)
                    )
                ):
                    missing_sips.append(sip.name)

                    continue

            sip_widget = SIPWidget(sip=sip)

            try:
                if sip.metadata_file_path != "":
                    sip_widget.metadata_df = pd.read_excel(
                        sip.metadata_file_path, dtype=str
                    )
            except Exception:
                missing_sips.append(sip.name)
                continue

            sip.value_changed.connect(self.state.update_sip)

            # Uploading is not a valid state, could have happened because of forced shutdown during upload
            if sip.status == SIPStatus.UPLOADING:
                sip.set_status(SIPStatus.SIP_CREATED)

            result = FileController.existing_grid(
                self.application.state.configuration, sip
            )

            if result is not None:
                grid = result

                sip_widget.import_template_df = grid
                sip_widget.import_template_location = os.path.join(
                    self.application.state.configuration.misc.save_location,
                    FileController.IMPORT_TEMPLATE_STORAGE,
                    f"{sip.series._id}.xlsx",
                )

            if sip.status != SIPStatus.IN_PROGRESS:
                sip_widget.open_button.setEnabled(False)

            if sip.status == SIPStatus.SIP_CREATED:
                sip_widget.upload_button.setEnabled(True)
                
            if sip.status in (
                SIPStatus.UPLOADED,
                SIPStatus.PROCESSING,
                SIPStatus.ACCEPTED,
                SIPStatus.REJECTED,
                SIPStatus.SIP_CREATED,
            ):
                sip_widget.open_explorer_button.setEnabled(True)
                sip_widget.delete_button.setEnabled(True)

            if sip.status in (SIPStatus.PROCESSING, SIPStatus.ACCEPTED, SIPStatus.REJECTED):
                sip_widget.open_edepot_button.setEnabled(True)

            self.sip_list_view.add_item(
                searchable_name_field="sip_name",
                widget=sip_widget,
            )

        if len(missing_sips) > 0:
            WarningDialog(
                title="Missende bestanden",
                text=f"Een of meerdere sips, sidecars of metadata zijn niet aanwezig.\n\nMissende sips: {json.dumps(missing_sips, indent=4)}\n\nDeze bestanden zijn nodig om gegevens in te laden, deze sips worden overgeslagen.",
            ).exec()

    def add_dossier_clicked(self, multi=False):
        dossier_path = QtWidgets.QFileDialog.getExistingDirectory(
            caption="Selecteer dossier om toe te voegen"
        )

        if dossier_path != "":
            paths = [dossier_path]

            if multi:
                paths = os.listdir(dossier_path)

            overlapping_labels = self.dossiers_list_view.get_overlapping_values(paths)

            unique_paths = [p for p in paths if p not in overlapping_labels]

            bad_dossiers = [
                os.path.normpath(os.path.join(dossier_path, partial_path))
                for partial_path in overlapping_labels
            ]
            dossiers = []
            dossier_widgets = []

            estimated_seconds = len(unique_paths) // 800

            if estimated_seconds > 2:
                WarningDialog(
                    title="Dossiers toevoegen",
                    text=f"Het toevoegen van veel dossiers kan een tijdje duren.\n\nGeschatte tijd: {estimated_seconds} seconden",
                ).exec()

            for partial_path in unique_paths:
                path = os.path.normpath(os.path.join(dossier_path, partial_path))

                # NOTE: we do not care about files in there, we only take the folders as dossiers
                if not os.path.isdir(path):
                    continue

                dossier = Dossier(path=path)
                dossiers.append(dossier)

                dossier_widget = DossierWidget(dossier=dossier)

                dossier_widgets.append(dossier_widget)

            self.dossiers_list_view.add_items(
                widgets=dossier_widgets,
                selection_changed_callback=self.dossier_selection_changed,
            )

            self.state.add_dossiers(dossiers=dossiers)

            if len(bad_dossiers) > 0:
                WarningDialog(
                    title="Dossiers niet toegevoegd",
                    text=f"Sommige dossiers overlappen in naamgeving met bestaande dossiers.\n\nDossiers die overlappen: {json.dumps(bad_dossiers, indent=4)}.\n\nVerander de namen van de dossiers (foldernamen) zodat ze uniek zijn in de lijst van dossiers en voeg opnieuw toe.",
                ).exec()

    def create_sip_clicked(self):
        selected_dossiers = list(self.dossiers_list_view.get_selected())

        if len(selected_dossiers) > 0:
            dossiers = [d.dossier for d in selected_dossiers]

            sip = SIP(
                environment_name=self.application.state.configuration.active_environment_name,
                dossiers=dossiers,
            )
            sip.value_changed.connect(self.state.update_sip)
            sip_widget = SIPWidget(sip=sip)

            success = self.sip_list_view.add_item(
                searchable_name_field="sip_id",
                widget=sip_widget,
            )

            if success:
                self.application.state.add_sip(sip)

                # Remove the dossiers from the list
                self.dossiers_list_view.remove_selected_clicked()

                # Open the SIP
                sip_widget.open_button_clicked()

    def dossier_selection_changed(self):
        self.create_sip_button.setEnabled(
            len(self.dossiers_list_view.get_selected()) > 0
        )

class MigrationWidget(QtWidgets.QWidget):
    def __init__(self, parent: MainWindow):
        super().__init__(parent)

        self.application: Application = QtWidgets.QApplication.instance()
        self.state: State = self.application.state
        
        self.list_storage_path = f"{self.state.configuration.misc.save_location}/overdrachtslijsten"

        self._layout = QtWidgets.QGridLayout()
        self.list_view = SearchableListWidget()

        self.main_db = "main.db"

    def setup_ui(self):
        self.setLayout(self._layout)

        # MAIN UI
        font = QtGui.QFont()
        font.setBold(True)
        font.setPointSize(20)

        title = QtWidgets.QLabel(text="Overdrachtslijsten")
        title.setFont(font)

        self.add_item_button = QtWidgets.QPushButton(text="Importeer overdrachtslijst")
        self.add_item_button.clicked.connect(self.add_overdrachtslijst_click)
        self.add_item_button.setHidden(self.state.configuration.active_role == "klant")
        self.parent().toolbar.configuration_changed.connect(self.hide_add_button)

        file_location_button = QtWidgets.QPushButton(text="Bestandslocatie")
        file_location_button.clicked.connect(lambda: os.startfile(
                os.path.join(
                    self.state.configuration.misc.save_location,
                    "overdrachtslijsten"
                )
            )
        )

        self._layout.addWidget(title, 0, 0)
        self._layout.addWidget(self.add_item_button, 1, 0)
        self._layout.addWidget(file_location_button, 1, 3)
        self._layout.addWidget(self.list_view, 2, 0, 1, 4)

        from creator.controllers.api_controller import APIController

        # NOTE: preload the series
        self.series = APIController.get_series(self.state.configuration)

    def load_items(self):
        os.makedirs(self.list_storage_path, exist_ok=True)

        for partial_path in os.listdir(self.list_storage_path):
            path = os.path.join(self.list_storage_path, partial_path)

            tab_ui = TabUI(path=path, series=self.series)
            tab_ui.setup_ui()
            tab_ui.load_items()

            self.list_view.add_item("overdrachtslijst_name", ListView(tab_ui))

    def add_overdrachtslijst_click(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            caption="Selecteer Overdrachtslijst", filter="Overdrachtslijst (*.xlsx *.xlsm *.xltx *.xltm)"
        )

        if path == "":
            return

        tab_ui = TabUI(path=path, series=self.series)
        
        if tab_ui.overdrachtslijst_name not in [w['reference'].overdrachtslijst_name for w in self.list_view.widgets]:
            tab_ui.setup_ui()
            tab_ui.load_items()
            self.list_view.add_item("overdrachtslijst_name", ListView(tab_ui))

        tab_ui.show()

    def hide_add_button(self) -> None:
        try:
            self.add_item_button.setHidden(self.state.configuration.active_role == "klant")
        except RuntimeError:
            # TODO: why does this happen? is there a cleaner solution?
            # NOTE: happens if the button already no longer exists
            pass


class TabUI(QtWidgets.QMainWindow):
    can_upload_changed: QtCore.Signal = QtCore.Signal(*(bool,), arguments=["can_upload"])
    edepot_available_changed: QtCore.Signal = QtCore.Signal(*(bool,), arguments=["edepot_available"])
    configuration_changed: QtCore.Signal = QtCore.Signal()

    def __init__(self, path: str, series: list):
        super().__init__()

        self.application: Application = QtWidgets.QApplication.instance()
        self.state: State = self.application.state

        self.can_upload = False
        self.edepot_ids = []

        self.storage_base = f"{self.state.configuration.misc.save_location}/overdrachtslijsten"

        self.toolbar = Toolbar()
        self.toolbar.configuration_changed.connect(self.configuration_changed.emit)

        self.path = path
        self.overdrachtslijst_name = Path(path).stem
        self.db_location = f"{self.storage_base}/{self.overdrachtslijst_name}.db"
        self.series = series

        self._layout = QtWidgets.QGridLayout()
        self.tabs: dict[str, TableView] = dict()

        self.main_tab = "Overdrachtslijst"
        self.main_table = TableView()

        self.tab_widget = QtWidgets.QTabWidget()

        self.configuration_changed.connect(self.reload_tabs)

    def setup_ui(self):
        self.resize(800, 600)
        self.setWindowTitle(self.overdrachtslijst_name)
        self.addToolBar(self.toolbar)

        central_widget = QtWidgets.QWidget()
        central_widget.setLayout(self._layout)
        self.setCentralWidget(central_widget)

        self.create_sips_button = QtWidgets.QPushButton(text="Maak SIPs")
        self.create_sips_button.clicked.connect(self.create_sips)
        self.create_sips_button.setHidden(self.state.configuration.active_role == "klant")
        self.configuration_changed.connect(lambda: self.hide_or_show_button(self.create_sips_button))
        self.configuration_changed.connect(self.set_create_button_status)

        self._layout.addWidget(self.tab_widget, 0, 0)

        save_button = QtWidgets.QPushButton(text="Opslaan")
        save_button.clicked.connect(self.save_tabs)
        self._layout.addWidget(save_button, 1, 0)
        self._layout.addWidget(self.create_sips_button, 2, 0)

    def load_items(self):
        if self.create_db():
            # No need to load if the db already existed

            try:
                self.load_overdrachtslijst()
            except:
                # NOTE: if an exception occurred and we just created the db, remove it again
                os.remove(self.db_location)
                raise
        
        self.load_main_tab()
        self.load_other_tabs()

        self.set_create_button_status()

        with sql.connect(self.db_location) as conn:
            result = conn.execute(f'''
                SELECT uploaded, edepot_id FROM tables;
            ''').fetchall()

        has_uploaded = False

        for uploaded, edepot_id in result:
            if edepot_id is not None:
                self.edepot_ids.append(edepot_id)

            has_uploaded = has_uploaded or uploaded

        if has_uploaded:
            self.can_upload = False
            self.can_upload_changed.emit(False)

            # NOTE: this means some of the items had not been found in the edepot yet
            if len(result) - 1 != len(self.edepot_ids):
                self.edepot_ids = []
                self.edepot_available_changed.emit(False)

                t = threading.Thread(
                    target=self.update_status
                )
                t.start()

                Dialog(
                    title="Zoeken E-depot",
                    text="Sommige items waren al geupload maar nog niet teruggevonden in het E-depot, deze worden nu verder gezocht.\n\nWanneer de link naar het E-depot beschikbaar is, zal de knop hiervoor actief worden."
                ).exec()
            else:
                self.edepot_available_changed.emit(True)

    def create_db(self) -> bool:
        import sqlite3 as sql
        import os

        os.makedirs(self.storage_base, exist_ok=True)


        if os.path.exists(self.db_location):
            with sql.connect(self.db_location) as conn:
                # ALTERS
                columns = conn.execute("PRAGMA table_info(tables);").fetchall()
                column_names = [column_name for _, column_name, *_ in columns]

                if 'uploaded' not in column_names:
                    conn.execute("""
                        ALTER TABLE tables
                        ADD COLUMN uploaded BOOLEAN;
                    """)

                    conn.execute("""
                        UPDATE tables
                        SET uploaded = CASE
                            WHEN edepot_id IS NOT NULL THEN 1
                            ELSE 0
                        END;
                    """)
            
                if 'URI Serieregister' not in column_names:
                    if 'uri_serieregister' in column_names:
                        conn.execute('ALTER TABLE tables RENAME COLUMN uri_serieregister TO "URI Serieregister";')
                    else:
                        conn.execute('ALTER TABLE tables ADD COLUMN "URI Serieregister" TEXT;')
            
            return False

        with sql.connect(self.db_location) as conn:
            conn.execute("""
            CREATE TABLE tables (
                id INTEGER PRIMARY KEY,
                table_name TEXT,
                "URI Serieregister" TEXT,
                edepot_id TEXT,
                uploaded BOOLEAN DEFAULT 0,

                UNIQUE(table_name)
            );""")

            conn.execute(f"""
            INSERT INTO tables (table_name)
            VALUES ('{self.main_tab}');
            """)

            conn.commit()

        return True

    def load_overdrachtslijst(self):
        import pandas as pd
        import sqlite3 as sql

        wb = load_workbook(
            self.path,
            read_only=True,
            data_only=True,
            keep_links=False,
            rich_text=False,
        )

        if not self.main_tab in wb.sheetnames:
            raise ValueError(f"{self.main_tab} tab missing")

        ws = wb[self.main_tab]
        data = ws.values

        try:
            # TODO: check all required columns?
            while "Doosnr" not in (headers := next(data)):
            # TODO: temp
            # while "Doosnr. " not in (headers := next(data)):
                pass
        except StopIteration:
            # TODO: proper error here
            raise Exception("Geen hoofdingen gevonden in de overdrachtslijst")

        expected_headers = (
            "Beschrijving",
            "Begindatum",
            "Einddatum",
            "Doosnr",
            "URI Serieregister",
        )
        # TODO: temp
        # expected_headers = {
        #     "Beschrijving ": "Beschrijving",
        #     "Begin-\ndatum": "Begindatum",
        #     "Eind-\ndatum": "Einddatum",
        #     "Doosnr. ": "Doosnr",
        #     None: "URI Serieregister",
        # }
        headers = [h.strip() for h in headers if h is not None]
        for h in expected_headers:
            if h not in headers:
                raise Exception(f"Verwachtte om de kolom '{h}' tegen te komen, maar is niet gevonden.")

        # TODO: temp
        # actual_headers = []

        # for h in headers:
        #     if h in expected_headers:
        #         actual_headers.append(expected_headers[h])
        #     else:
        #         actual_headers.append(h)

        # Filter out empty rows
        df = pd.DataFrame(
            (
                r for r in 
                (r[:len(headers)] for r in list(data))
                if not all(not bool(v) for v in r)
            ),
            # TODO: temp
            columns=headers,
            # columns=actual_headers,
        ).fillna("").astype(str).convert_dtypes()
        wb.close()

        # TODO: temp
        # df["URI Serieregister"] = "https://serieregister-ti.vlaanderen.be/id/serie/e641d8943266475594d43bd7e9d9bb08ea4893ce5e9646e39bc56911bfffc079"
        df["id"] = range(df.shape[0])
        df["series_name"] = ""

        # NOTE: since Excel deals with datetimes awkwardly, make sure we only have the date part as a string here
        # Only take the date part, if none found, keep original value
        df["Begindatum"] = df["Begindatum"].str.extract(r"(\d{4}-\d{2}-\d{2})", expand=False).fillna(df["Begindatum"])
        df["Einddatum"] = df["Einddatum"].str.extract(r"(\d{4}-\d{2}-\d{2})", expand=False).fillna(df["Einddatum"])

        # NOTE: reorder headers
        cols = df.columns.tolist()
        cols = ["id", "series_name", *(c for c in cols if c not in ("id", "series_name", "URI Serieregister")), "URI Serieregister"]
        df = df[cols]

        con = sql.connect(self.db_location)
        df.to_sql(
            name=self.main_tab,
            con=con,
            index=False,
            method="multi",
            # if_exists="append",
            chunksize=1000,
        )

    def load_main_tab(self):
        container = QtWidgets.QWidget()
        layout = QtWidgets.QGridLayout()
        container.setLayout(layout)

        listed_series = self.series
        series_names = [s.get_name() for s in listed_series if s.status == "Published"]
        
        series_combobox = QtWidgets.QComboBox()
        series_combobox.setEditable(True)
        series_combobox.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        series_combobox.completer().setCompletionMode(
            QtWidgets.QCompleter.PopupCompletion
        )
        series_combobox.completer().setFilterMode(
            QtCore.Qt.MatchFlag.MatchContains
        )
        series_combobox.setMaximumWidth(900)
        series_combobox.addItems(series_names)

        btn = QtWidgets.QPushButton(text="Voeg toe")
        btn.clicked.connect(
            lambda: 
            self.add_to_new(
                name = series_combobox.currentText(),
                series_id = listed_series[series_names.index(series_combobox.currentText())]._id
            )
        )

        model = SQLliteModel(self.main_tab, db_name=self.db_location, is_main=True)
        model.bad_rows_changed.connect(self.set_create_button_status)
        self.main_table.setModel(model)
        
        self.unassigned_only_checkbox = QtWidgets.QCheckBox(text="Toon enkel rijen zonder serie")
        self.unassigned_only_checkbox.stateChanged.connect(self._filter_unassigned)

        layout.addWidget(btn, 0, 0)
        layout.addWidget(series_combobox, 0, 1, 1, 3)
        layout.addWidget(self.unassigned_only_checkbox, 1, 0)
        layout.addWidget(self.main_table, 2, 0, 1, 5)

        self.tab_widget.addTab(container, self.main_tab)
        self.tabs[self.main_tab] = self.main_table

        # NOTE: hide the id column
        self.main_table.hideColumn(0)

        # NOTE: map all the URI Serieregisters
        conn = sql.connect(self.db_location)
        with conn:
            unique_serie_uris = [
                uri for uri, *_ in 
                conn.execute(f"""SELECT "URI Serieregister" FROM {self.main_tab} GROUP BY "URI Serieregister";""").fetchall()
            ]

            uri_index_maps = {}

            for serie_uri in unique_serie_uris:
                uri_index_maps[serie_uri] = [
                    i for i, *_ in
                    conn.execute(f"""SELECT id FROM {self.main_tab} WHERE "URI Serieregister" = '{serie_uri}';""").fetchall()
                ]

        # NOTE: set all the series_names where the series_id matches one we got
        uri_pre = self.state.configuration.active_environment.get_serie_register_uri()

        for uri, indexes in uri_index_maps.items():
            match = [s for s in self.series if s._id == uri.split(uri_pre + "/")[-1]]

            if len(match) != 1:
                continue

            series = match[0]

            self.add_to_new(name=series.get_name(), series_id=series._id, mapping_ids=indexes)

    def load_other_tabs(self):
        conn = sql.connect(self.db_location)

        with conn:
            tables = conn.execute(f"""SELECT table_name, "URI Serieregister" FROM tables WHERE table_name != '{self.main_tab}';""")
        
        # NOTE: set all the series_names where the series_id matches one we got
        uri_pre = self.state.configuration.active_environment.get_serie_register_uri()

        for table_name, uri_serieregister in tables:
            match = [s for s in self.series if s._id == uri_serieregister.split(uri_pre + "/")[-1]]

            if len(match) != 1:
                continue

            series = match[0]

            # NOTE: remove leading and trailing quotes
            self.create_tab(name=table_name[1:-1], series_id=series._id)

    def add_to_new(self, name: str, series_id: str, mapping_ids: list[int] = None):
        # NOTE: only thing not allowed is quotes
        name = name.strip().replace('"', "").replace("'", "")

        # No funny business
        if name == "" or name == self.main_tab:
            return

        conn = sql.connect(self.db_location)

        if mapping_ids is None:
            selected_rows = [str(r.row()) for r in self.main_table.selectionModel().selectedRows()]

            if len(selected_rows) == 0:
                return

            selected_rows_str = ", ".join(selected_rows)
        else:
            selected_rows_str = ", ".join(str(i) for i in mapping_ids)

        uri = f"{self.state.configuration.active_environment.get_serie_register_uri()}/{series_id}"

        with conn:
            # Check if table exists
            result = conn.execute(f'pragma table_info("{name}");').fetchall()

            # Create table
            if not result:
                import_sjabloon = APIController.get_import_template(self.state.configuration, series_id=series_id)

                columns = pd.read_excel(import_sjabloon, dtype=str, engine="openpyxl").columns

                conn.execute(f"""
                CREATE TABLE IF NOT EXISTS "{name}" (
                    id INTEGER PRIMARY KEY,

                    main_id INTEGER NOT NULL,

                    {",\n\t".join(f'"{c}" TEXT' for c in columns)}
                );""")

                # Update the tables table
                conn.execute(f"""
                    INSERT OR IGNORE INTO tables (table_name, "URI Serieregister")
                    VALUES ('"{name}"', '{uri}');
                """)

            # Remove where needed
            cursor = conn.execute(f"""
                SELECT id, series_name
                FROM {self.main_tab}
                WHERE id IN ({selected_rows_str})
                  AND series_name != '"{name}"'
                  AND series_name != '';
            """)

            for main_id, table in cursor.fetchall():
                # NOTE: table already contains ""-marks
                tab = table[1:-1]

                conn.execute(f"""
                    DELETE FROM {table}
                    WHERE main_id={main_id};
                """)

                rows = conn.execute(f"""
                    SELECT count() FROM {table};
                """).fetchone()[0]

                if rows == 0:
                    conn.execute(f"""
                        DROP TABLE {table};
                    """)
                    
                    conn.execute(f"""
                        DELETE FROM tables
                        WHERE table_name='{table}';
                    """)

                    self.tab_widget.removeTab(list(self.tabs).index(tab))
                    del self.tabs[tab]
                    
                    continue

                # Recalculate shape for table
                model: SQLliteModel = self.tabs[tab].model()
                model.row_count = rows

                # Update the graphical side
                model.layoutChanged.emit()

            # Insert where needed
            # NOTE: don't do other auto-mapping
            conn.execute(f"""
                INSERT INTO "{name}" (main_id, "Analoog?", "Path in SIP", "DossierRef", "Naam", "Openingsdatum", "Sluitingsdatum", "Origineel Doosnummer")
                SELECT id, 'ja', "Beschrijving", "Beschrijving", "Beschrijving", "Begindatum", "Einddatum", substr('0000' || "Doosnr", -4, 4) || '/{self.overdrachtslijst_name}'
                FROM {self.main_tab}
                WHERE id IN ({selected_rows_str})
                  AND (series_name != '"{name}"' OR series_name IS NULL OR series_name == '');
            """)
            
            # Update the main table to show correct linking
            conn.execute(f"""
                UPDATE {self.main_tab}
                SET series_name='"{name}"',
                    "URI Serieregister"='{uri}'
                WHERE id IN ({selected_rows_str});
            """)

            conn.commit()
        
        # If the tab already exists, just update the data
        if name in self.tabs:
            model: SQLliteModel = self.tabs[name].model()
            
            model.get_data()
            model.layoutChanged.emit()
        elif mapping_ids:
            # NOTE: we added this automatically, don't add the tab here
            pass
        else:
            self.create_tab(name, series_id)

        # Update the graphical side for all tables involved
        model: SQLliteModel = self.main_table.model()
        
        model.get_data()
        model.layoutChanged.emit()

        for table in self.tabs.values():
            model: SQLliteModel = table.model()

            model.get_data()
            model.layoutChanged.emit()

        self._filter_unassigned(self.unassigned_only_checkbox.checkState().value)

    def create_tab(self, name: str, series_id: str):
        from creator.widgets.tableview_widget import TableView

        container = QtWidgets.QWidget()
        layout = QtWidgets.QGridLayout()
        container.setLayout(layout)

        series_label = QtWidgets.QLabel(text=name)
        
        duplicate_trefwoord_column_button = QtWidgets.QPushButton(text="Dupliceer trefwoorden_vrij kolom")
        duplicate_location_column_button = QtWidgets.QPushButton(text="Dupliceer locatie kolommen")
        duplicate_location_column_button.setHidden(self.state.configuration.active_role == "klant")
        self.configuration_changed.connect(lambda: self.hide_or_show_button(duplicate_location_column_button))

        table_view = TableView()
        bad_rows_checkbox = QtWidgets.QCheckBox(text="Toon enkel rijen met fouten")

        layout.addWidget(series_label, 0, 0, 1, 2)
        layout.addWidget(duplicate_trefwoord_column_button, 1, 0)
        layout.addWidget(duplicate_location_column_button, 1, 1)
        layout.addWidget(bad_rows_checkbox, 2, 0)
        layout.addWidget(table_view, 3, 0, 1, 5)

        model = SQLliteModel(name, db_name=self.db_location, series_id=series_id)
        table_view.setModel(model)
        
        load_bestandscontrole_button = QtWidgets.QPushButton(text="Laad bestandscontrole lijst")
        load_bestandscontrole_button.clicked.connect(lambda _: self.load_bestandscontrole(model=model))
        load_bestandscontrole_button.setHidden(self.state.configuration.active_role == "klant")
        self.configuration_changed.connect(lambda: self.hide_or_show_button(load_bestandscontrole_button))
        layout.addWidget(load_bestandscontrole_button, 1, 4)

        bad_rows_checkbox.stateChanged.connect(lambda checkstate: self._filter_bad_rows(checkstate, self.tabs[name]))
        model.bad_rows_changed.connect(lambda _: self._filter_bad_rows(bad_rows_checkbox.checkState().value, table_view))
        model.layoutChanged.connect(lambda _: self._filter_bad_rows(bad_rows_checkbox.checkState().value, table_view))
        model.bad_rows_changed.connect(self.set_create_button_status)
        self.configuration_changed.connect(lambda: self._filter_bad_rows(bad_rows_checkbox.checkState().value, table_view))

        self.tab_widget.addTab(container, name)
        self.tabs[name] = table_view

        duplicate_trefwoord_column_button.clicked.connect(lambda _: self.add_column(name))
        duplicate_location_column_button.clicked.connect(lambda _: self.add_column(name, location_cols=True))

        # NOTE: hide id and main_id
        table_view.hideColumn(0)
        table_view.hideColumn(1)

        if self.state.configuration.active_role == "klant":
            cols_to_skip = ("Origineel Doosnummer", "Legacy locatie ID", "Legacy range", "Verpakkingstype")

            with sql.connect(self.db_location) as conn:
                # NOTE: figure out which columns to hide (could be multiple due to duplications)
                
                cursor = conn.execute(f"pragma table_info(\"{name}\");")

                for i, column_name, *_ in cursor.fetchall():
                    if any(c in column_name for c in cols_to_skip):
                        table_view.hideColumn(i)

    def closeEvent(self, event):
        from creator.utils.sqlitemodel import SQLliteModel

        models: list[SQLliteModel] = [t.model() for t in self.tabs.values()]

        # NOTE: only ask the user if data has actually changed
        if not any(m.has_changed for m in models):
            event.accept()
            return

        dialog = YesNoDialog(
            title="Overdrachtslijst sluiten",
            text="Wil je de huidige data opslaan?\nZonder opslaan gaat de nieuwe data verloren bij heropstarten van de applicatie."
        )
        dialog.exec()

        if dialog.result():
            self.save_tabs()

        event.accept()

    def save_tabs(self) -> None:
        from creator.utils.sqlitemodel import SQLliteModel

        for table_view in self.tabs.values():
            if table_view == self.main_table:
                continue

            model: SQLliteModel = table_view.model()

            model.save_data()
            model.has_changed = False

        self.close()

    def reload_tabs(self) -> None:
        for table_name, tab_view in self.tabs.items():
            # Reload all the data
            model: SQLliteModel = tab_view.model()
            model.get_data()

            with sql.connect(self.db_location) as conn:
                # NOTE: figure out which columns to hide (could be multiple due to duplications)
                cursor = conn.execute(f"pragma table_info(\"{table_name}\");")

                columns = cursor.fetchall()

            # Show every column
            for i in range(len(columns)):
                tab_view.showColumn(i)

            # Hide id and main_id columns where applicable
            for i, column_name, *_ in columns:
                if column_name in ("id", "main_id"):
                    tab_view.hideColumn(i)

            if self.state.configuration.active_role == "klant":
                cols_to_skip = ("Origineel Doosnummer", "Legacy locatie ID", "Legacy range", "Verpakkingstype")

                for i, column_name, *_ in columns:
                    if any(c in column_name for c in cols_to_skip):
                        tab_view.hideColumn(i)

    def set_create_button_status(self, *_) -> None:
        for table_view in self.tabs.values():
            model: SQLliteModel = table_view.model()
            red_colors = [_ for c in model.colors.values() if c == Color.RED]

            if len(red_colors) > 0:
                self.create_sips_button.setEnabled(False)
                self.can_upload = False
                self.can_upload_changed.emit(False)
                return

        self.create_sips_button.setEnabled(True)
        self.can_upload = True
        self.can_upload_changed.emit(True)

    def hide_or_show_button(self, button: QtWidgets.QPushButton) -> None:
        button.setHidden(self.state.configuration.active_role == "klant")

    def add_column(self, table: str, location_cols: bool=False) -> None:
        def find_next(old_columns: list, name: str) -> tuple[str, str]:
            # Returns new column name, and the previous one to look for
            value = 1

            while (new_name := f'{name}_{value}') in old_columns:
                value += 1

            return new_name, f'{name}_{value-1}' if value > 1 else name

        with sql.connect(self.db_location) as conn:
            # NOTE: since we need to reset the whole table definition
            # Rename the table
            conn.execute(f'ALTER TABLE "{table}" RENAME TO "_old_{table}";')

            # Get old table definition
            cursor = conn.execute(f'pragma table_info("_old_{table}");')

            old_definitions = [(name, _type) for _, name, _type, *_ in cursor.fetchall()]
            old_columns = [d[0] for d in old_definitions]

            # Find where to input the new column
            new_column_name, previous_column_name = find_next(old_columns, "Origineel Doosnummer" if location_cols else "Trefwoorden_vrij")
            
            if location_cols:
                previous_column_name = previous_column_name.replace("Origineel Doosnummer", "Verpakkingstype")

            # Create the new definitions
            new_definitions = []

            for column_name, column_type in old_definitions:
                new_definitions.append(f'"{column_name}" {column_type}')

                if column_name == previous_column_name:
                    if location_cols:
                        columns = ("Origineel Doosnummer", "Legacy locatie ID", "Legacy range", "Verpakkingstype")
                        number = new_column_name.split('_')[-1]

                        for c in columns:
                            new_definitions.append(f'"{c}_{number}" {column_type}')
                    else:
                        new_definitions.append(f'"{new_column_name}" {column_type}')

            # Create new table
            conn.execute(f'''
                CREATE TABLE "{table}" (
                    {",\n\t".join(new_definitions)}
                );
            ''')

            # Fill table with old values
            conn.execute(f'''
                INSERT INTO "{table}" ({", ".join(f'"{c}"' for c in old_columns)})
                SELECT {", ".join(f'"{c}"' for c in old_columns)}
                FROM "_old_{table}";
            ''')

            conn.execute(f'DROP TABLE "_old_{table}";')

        model: SQLliteModel = self.tabs[table].model()
        model.get_data()
        model.layoutChanged.emit()

    def _filter_unassigned(self, state: QtCore.Qt.CheckState) -> None:
        model: SQLliteModel = self.main_table.model()
        data: list[list[str]] = model.raw_data

        if state == QtCore.Qt.CheckState.Checked.value:
            columns = model.columns
            series_col = list(columns.values()).index("series_name")

            for row_index, row in enumerate(data):
                if row[series_col] != "":
                    self.main_table.setRowHidden(row_index, True)

            return
        
        for row_index in range(len(data)):
            self.main_table.setRowHidden(row_index, False)

    def _filter_bad_rows(self, state: QtCore.Qt.CheckState, table_view: TableView) -> None:
        model: SQLliteModel = table_view.model()

        if state != QtCore.Qt.CheckState.Checked.value:
            # NOTE: show all
            for i in range(model.row_count):
                table_view.showRow(i)

            return

        # NOTE: since sometimes columns might be hidden, we need to make sure we skip those
        ids_to_show = set(_id for (_id, col_index) in model.colors.keys() if not table_view.isColumnHidden(col_index))

        for i, row in enumerate(model.raw_data):
            _id = int(row[0])

            if _id in ids_to_show:
                table_view.showRow(i)
            else:
                table_view.hideRow(i)

    def create_sips(self) -> None:
        def _col_index_to_xslx_col(col_index: int) -> str:
            # NOTE: this only supports up to ZZ for now
            first_letter_value = (col_index // 26) - 1

            if first_letter_value >= 26:
                raise ValueError("There are too many columns")
            if first_letter_value == -1:
                return chr(65 + col_index)

            first_letter = chr(65 + first_letter_value)
            second_letter = chr(65 + col_index % 26)
            
            return f"{first_letter}{second_letter}"

        storage_location = self.state.configuration.misc.save_location
        sjabloon_base_path = os.path.join(storage_location, FileController.IMPORT_TEMPLATE_STORAGE)
        sip_storage_path = os.path.join(storage_location, FileController.SIP_STORAGE)
        grid_storage_path = os.path.join(storage_location, FileController.GRID_STORAGE)

        os.makedirs(sip_storage_path, exist_ok=True)
        os.makedirs(grid_storage_path, exist_ok=True)

        for series_name, table_view in self.tabs.items():
            if series_name == self.main_tab:
                continue

            model: SQLliteModel = table_view.model()
            model.save_data()

            # Copy import_template to grid_storage
            temp_loc = os.path.join(grid_storage_path, f"temp_{model.series_id}.xlsx")

            wb = load_workbook(os.path.join(sjabloon_base_path, f"{model.series_id}.xlsx"))
            ws = wb["Details"]

            data = model.raw_data

            # NOTE: overwrite the headers
            for col_index, col in model.columns.items():
                # NOTE: skip id and main_id
                if col_index < 2:
                    continue

                # NOTE: duplicate columns generate like <col>_<n>
                # duplicate columns read generate like <col>.<n>
                re_match = re.match(r"(.*)[_.]\d+$", col)

                if re_match:
                    col = re_match.group(1)

                ws[f"{_col_index_to_xslx_col(col_index - 2)}1"] = col

            for row_index, row in enumerate(data):
                for col_index, value in enumerate(row):
                    # NOTE: skip id and main_id
                    if col_index < 2:
                        continue

                    ws[f"{_col_index_to_xslx_col(col_index - 2)}{row_index+2}"] = value

            wb.save(temp_loc)
            wb.close()

            sip_location = os.path.join(sip_storage_path, f"{model.series_id}-{self.overdrachtslijst_name}.zip")
            md5_location = os.path.join(sip_storage_path, f"{model.series_id}-{self.overdrachtslijst_name}.xml")

            with zipfile.ZipFile(
                sip_location, "w", compression=zipfile.ZIP_DEFLATED
            ) as zfile:
                zfile.write(
                    temp_loc,
                    "Metadata.xlsx"
                )

            md5 = hashlib.md5(open(sip_location, "rb").read()).hexdigest()

            side_car_info = """<?xml version="1.0" encoding="UTF-8"?>
<mhs:Sidecar xmlns:mhs="https://zeticon.mediahaven.com/metadata/20.3/mhs/" version="20.3" xmlns:mh="https://zeticon.mediahaven.com/metadata/20.3/mh/">
    <mhs:Technical>
            <mh:Md5>{md5}</mh:Md5>
    </mhs:Technical>
</mhs:Sidecar>""".format(
                md5=md5
            )

            with open(md5_location, "w", encoding="utf-8") as f:
                f.write(side_car_info)

            os.remove(temp_loc)
        
            # NOTE; set the edepot_id to empty again, since it hasn't been uploaded in it's current form
            with sql.connect(self.db_location) as conn:
                conn.execute(f'''
                    UPDATE tables
                    SET edepot_id=null,
                        uploaded=0
                    WHERE table_name='"{series_name}"';
                ''')

        self.can_upload = True
        self.can_upload_changed.emit(True)
        self.edepot_available_changed.emit(False)

        Dialog(
            title="SIPs aangemaakt",
            text="SIPS zijn aangemaakt voor de overdrachtslijst."
        ).exec()

        self.close()

    def upload_sips(self) -> None:
        env = self.state.configuration.active_environment
        if not env.has_ftps_credentials():
            WarningDialog(
                title="Connectie fout",
                text=f"Je FTPS connectie gegevens staan niet in orde voor omgeving '{self.sip.environment.name}'",
            ).exec()
            return

        storage_location = self.state.configuration.misc.save_location
        sip_storage_path = os.path.join(storage_location, FileController.SIP_STORAGE)

        for series_name, table_view in self.tabs.items():
            if series_name == self.main_tab:
                continue

            model: SQLliteModel = table_view.model()

            sip_location = os.path.join(sip_storage_path, f"{model.series_id}-{self.overdrachtslijst_name}.zip")
            md5_location = os.path.join(sip_storage_path, f"{model.series_id}-{self.overdrachtslijst_name}.xml")

            if not os.path.exists(sip_location) or not os.path.exists(md5_location):
                self.create_sips()
            
            try:
                with ftplib.FTP_TLS(
                    env.ftps_url,
                    env.ftps_username,
                    env.ftps_password,
                ) as session:
                    session.prot_p()

                    with open(sip_location, "rb") as f:
                        session.storbinary(f"STOR {model.series_id}-{self.overdrachtslijst_name}.zip", f)
                    with open(md5_location, "rb") as f:
                        session.storbinary(f"STOR {model.series_id}-{self.overdrachtslijst_name}.xml", f)
            except ftplib.error_perm:
                WarningDialog(
                    title="Connectie fout",
                    text=f"Je FTPS connectie login gegevens staan niet in orde voor omgeving '{self.sip.environment.name}'",
                ).exec()
                return
            except socket.gaierror:
                WarningDialog(
                    title="Connectie fout",
                    text=f"Je FTPS connectie url staat niet in orde voor omgeving '{self.sip.environment.name}'",
                ).exec()
                return
 
            with sql.connect(self.db_location) as conn:
                conn.execute(f'''
                    UPDATE tables
                    SET uploaded=1
                    WHERE table_name='"{series_name}"';
                ''')

        t = threading.Thread(
            target=self.update_status
        )
        t.start()

        Dialog(
            title="Upload geslaagd",
            text="Upload voor de overdrachtslijst is geslaagd.\nDe overdrachtslijst blijft in de lijst staan zolang hij op je computer staat.\nOm hem weg te halen, verwijder de correcte database uit de bestandslocatie.\n\nWanneer de items op het E-depot staan zal de knop hiervoor beschikbaar worden.\n\nZodra de link naar het E-depot beschikbaar is, zal de knop ook actief worden."
        ).exec()
        self.can_upload_changed.emit(False)

    def update_status(self) -> None:
        import time

        for series_name, table_view in self.tabs.items():
            if series_name == self.main_tab:
                continue

            model: SQLliteModel = table_view.model()
            edepot_id = None
            times_slept = 0
            max_time_to_sleep = 300

            while edepot_id is None and times_slept < max_time_to_sleep:
                edepot_id = APIController.get_sip_id_for_name(
                    self.state.configuration.active_environment,
                    f"{model.series_id}-{self.overdrachtslijst_name}.zip"
                )

                # NOTE: wait some time for the edepot to pick them up
                time.sleep(10)

                times_slept += 10

            if times_slept == max_time_to_sleep:
                Dialog(
                    title=f"SIP niet binnen de {max_time_to_sleep // 60} minuten op het E-depot gevonden",
                    text="De SIP was succesvol opgeladen via FTP, maat is niet binnen de tijd opgepikt door het E-depot."
                ).exec()
                break

            with sql.connect(self.db_location) as conn:
                conn.execute(f'''
                    UPDATE tables
                    SET edepot_id='{edepot_id}'
                    WHERE table_name='"{series_name}"';
                ''')

                self.edepot_ids.append(edepot_id)

        self.edepot_available_changed.emit(True)

    def load_bestandscontrole(self, model: SQLliteModel) -> None:
        controller = self.application.bestands_controle_lijst_controller

        if not controller.valid:
            WarningDialog(
                title="Bestandscontrole lijst is niet geldig",
                text="De bestandscontrole lijst is niet geldig, bekijk of het pad juist staat, en/of het bestand in orde is."
            ).exec()
            return

        values = controller.get_values(overdrachtslijst_name=self.overdrachtslijst_name)

        if values is None:
            # NOTE: a warning has already been shown to the user
            return

        if model.row_count > 1000:
            WarningDialog(
                title="Trage actie",
                text="Omdat er veel rijen moeten veranderen kan de actie tot enkele minuten duren."
            ).exec()

        col_indeces = []

        for i, col in model.columns.items():
            if col in ("Legacy locatie ID", "Legacy range", "Verpakkingstype"):
                col_indeces.append(i)

                if col == "Legacy locatie ID":
                    new_val = values[controller.list_start_column]
                elif col == "Legacy range":
                    new_val = values[controller.list_end_column]
                elif col == "Verpakkingstype":
                    new_val = values[controller.doos_type_column]
                
                non_empty_val = new_val != ""

                for r in range(0, model.row_count):
                    model.set_value(model.index(r, i), new_value=new_val)

                    if non_empty_val:
                        # NOTE: manually unmark, since this is waaaaay faster
                        model._mark_cell(r, i)

        model.dataChanged.emit(
            model.index(0, min(col_indeces)),
            model.index(model.row_count, max(col_indeces))
        )


class ListView(QtWidgets.QWidget):
    def __init__(self, tab_ui: TabUI):
        super().__init__()

        self.application: Application = QtWidgets.QApplication.instance()
        self.state: State = self.application.state

        self.overdrachtslijst_name = tab_ui.overdrachtslijst_name
        self.tab_ui = tab_ui

        layout = QtWidgets.QGridLayout()
        self.setLayout(layout)

        title = QtWidgets.QLabel(text=self.overdrachtslijst_name)

        open_button = QtWidgets.QPushButton(text="Open")
        open_button.clicked.connect(self.tab_ui.show)

        self.upload_button = QtWidgets.QPushButton(text="Upload")
        self.upload_button.clicked.connect(self.tab_ui.upload_sips)
        self.upload_button.setHidden(self.state.configuration.active_role == "klant")
        self.tab_ui.configuration_changed.connect(lambda: self.upload_button.setHidden(self.state.configuration.active_role == "klant"))

        self.upload_button.setEnabled(self.tab_ui.can_upload)
        self.tab_ui.can_upload_changed.connect(self.upload_button.setEnabled)

        self.edepot_button = QtWidgets.QPushButton(text="Open E-depot")
        self.edepot_button.clicked.connect(
            lambda: [os.startfile(
                f"{self.state.configuration.active_environment.api_url}/input/processing-list/{edepot_id}"
            ) for edepot_id in self.tab_ui.edepot_ids]
        )
        self.edepot_button.setHidden(self.state.configuration.active_role == "klant")
        self.tab_ui.configuration_changed.connect(lambda: self.edepot_button.setHidden(self.state.configuration.active_role == "klant"))

        self.edepot_button.setEnabled(len(self.tab_ui.edepot_ids) == len(self.tab_ui.tabs) - 1 and len(self.tab_ui.edepot_ids) > 0)
        self.tab_ui.edepot_available_changed.connect(self.edepot_button.setEnabled)

        layout.addWidget(title, 0, 0, 1, 3)
        layout.addWidget(open_button, 0, 3)
        layout.addWidget(self.upload_button, 1, 3)
        layout.addWidget(self.edepot_button, 2, 3)


def set_main(application: Application, main: MainWindow) -> None:
    config = application.state.configuration

    # TODO: use role
    active_role, active_type = config.active_role, config.active_type

    if active_type == "digitaal":
        main.central_widget = DigitalWidget(main)
        main.setWindowTitle("SIP Creator digitaal")
    else:
        main.central_widget = MigrationWidget(main)
        main.setWindowTitle("SIP Creator migratie")

    main.setCentralWidget(None)
    main.setCentralWidget(main.central_widget)
    main.central_widget.setup_ui()
    main.central_widget.load_items()


