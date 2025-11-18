import os
import re
import threading
import time
from typing import Iterable

import zipfile
import hashlib
import ftplib
import socket
from openpyxl import load_workbook
from PySide6 import QtWidgets, QtCore, QtGui
from PySide6.QtCore import QObject, Signal, QThread, Slot
import pandas as pd
import sqlite3 as sql
from pathlib import Path

from creator.application import Application, natural_keys

from creator.widgets.main_widgets.main_widget import MainWidget
from creator.widgets.searchable_list_widget import SearchableListWidget
from creator.widgets.tableview_widget import TableView
from creator.widgets.toolbar import Toolbar
from creator.widgets.dialog import YesNoDialog, Dialog, ChoiceDialog
from creator.widgets.warning_dialog import WarningDialog

from creator.controllers.api_controller import APIController
from creator.controllers.file_controller import FileController

from creator.utils.state import State
from creator.utils.sqlitemodel import SQLliteModel, CellColor
from creator.utils.proxymodel import CustomSortFilterModel
from creator.utils.path_loader import resource_path

from creator.windows.mainwindow import MainWindow


class MigrationWidget(MainWidget):
    def __init__(self, parent: MainWindow):
        super().__init__(parent)

        self.application: Application = QtWidgets.QApplication.instance()
        self.state: State = self.application.state
        
        self.list_storage_path = self.state.configuration.overdrachtslijsten_location

        self._layout = QtWidgets.QGridLayout()
        self.list_view = SearchableListWidget()

    def setup_ui(self):
        self.setLayout(self._layout)

        # MAIN UI
        font = QtGui.QFont()
        font.setBold(True)
        font.setPointSize(20)

        title = QtWidgets.QLabel(text="Overdrachtslijsten")
        title.setFont(font)

        self.add_item_button = QtWidgets.QPushButton(text="Importeer overdrachtslijst")
        self.add_item_button.clicked.connect(self.add_overdrachtslijst_click)
        self.add_item_button.setHidden(self.state.configuration.active_role == "klant")
        self.parent().toolbar.configuration_changed.connect(self.hide_add_button)

        file_location_button = QtWidgets.QPushButton(text="Bestandslocatie")
        file_location_button.clicked.connect(lambda: os.startfile(self.list_storage_path))

        self._layout.addWidget(title, 0, 0)
        self._layout.addWidget(self.add_item_button, 1, 0)
        self._layout.addWidget(file_location_button, 1, 3)
        self._layout.addWidget(self.list_view, 2, 0, 1, 4)

        self.series = self.state.series

    def load_items(self):
        os.makedirs(self.list_storage_path, exist_ok=True)

        for partial_path in os.listdir(self.list_storage_path):
            path = os.path.join(self.list_storage_path, partial_path)

            tab_ui = TabUI(path=path)
            tab_ui.setup_ui()
            # Only setting up the UI, data loading starts when a user
            # opens an overdrachtslijst.

            self.list_view.add_item("overdrachtslijst_name", MigrationListView(tab_ui))

    def add_overdrachtslijst_click(self):
        if not self.state.check_series_loaded():
            return

        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            caption="Selecteer Overdrachtslijst", filter="Overdrachtslijst (*.xlsx *.xlsm *.xltx *.xltm)"
        )

        if path == "":
            return

        tab_ui = TabUI(path=path)
        
        if tab_ui.overdrachtslijst_name not in [w['reference'].overdrachtslijst_name for w in self.list_view.widgets]:
            tab_ui.setup_ui()
            tab_ui.load_items()
            self.list_view.add_item("overdrachtslijst_name", MigrationListView(tab_ui))

        tab_ui.show()

    def hide_add_button(self) -> None:
        try:
            self.add_item_button.setHidden(self.state.configuration.active_role == "klant")
        except RuntimeError:
            # TODO: why does this happen? is there a cleaner solution?
            # NOTE: happens if the button already no longer exists
            pass


class TabUI(QtWidgets.QMainWindow):
    can_upload_changed: QtCore.Signal = QtCore.Signal(*(bool,), arguments=["can_upload"])
    edepot_available_changed: QtCore.Signal = QtCore.Signal(*(bool,), arguments=["edepot_available"])
    configuration_changed: QtCore.Signal = QtCore.Signal()

    def __init__(self, path: str):
        super().__init__()

        self.application: Application = QtWidgets.QApplication.instance()
        self.state: State = self.application.state
        self.state.check_series_loaded()
        self.series = self.state.series

        self.setWindowIcon(QtGui.QIcon(resource_path("logo.ico")))

        self.can_upload = False
        self._loaded = False # A marker so the UI is not populated multiple times when a user closes and reopens an overdrachtslijst
        self.edepot_ids = []

        self.storage_base = self.state.configuration.overdrachtslijsten_location

        self.toolbar = Toolbar()
        self.toolbar.configuration_changed.connect(self.configuration_changed.emit)

        self.path = path
        self.overdrachtslijst_name = Path(path).stem
        self.db_location = f"{self.storage_base}/{self.overdrachtslijst_name}.db"

        self._layout = QtWidgets.QGridLayout()
        self.tabs: dict[str, [str, QtWidgets.QWidget]] = dict()

        # self.tabs look like this:
        # {
        #     <name>: {
        #         "container": <>,
        #         "table": <>,
        #     }
        # }

        self.main_tab = "Overdrachtslijst"
        self.main_table = TableView(editable=False)

        self.tab_widget = QtWidgets.QTabWidget()

        self.configuration_changed.connect(self.reload_tabs)

    def setup_ui(self):
        self.resize(800, 600)
        self.setWindowTitle(self.overdrachtslijst_name)
        self.addToolBar(self.toolbar)

        central_widget = QtWidgets.QWidget()
        central_widget.setLayout(self._layout)
        self.setCentralWidget(central_widget)

        self.create_sips_button = QtWidgets.QPushButton(text="Maak SIPs")
        self.create_sips_button.clicked.connect(self.create_sips)
        self.create_sips_button.setHidden(self.state.configuration.active_role == "klant")
        self.create_sips_button.setEnabled(False)
        self.configuration_changed.connect(lambda: self.hide_or_show_button(self.create_sips_button))
        self.configuration_changed.connect(self.set_create_button_status)

        self._layout.addWidget(self.tab_widget, 0, 0)

        self.save_button = QtWidgets.QPushButton(text="Opslaan")
        self.save_button.clicked.connect(self.save_tabs)
        self._layout.addWidget(self.save_button, 1, 0)
        self._layout.addWidget(self.create_sips_button, 2, 0)

        self._loaded = False

    def load_items(self):
        if self.create_db():
            # No need to load if the db already existed

            try:
                self.load_overdrachtslijst()
            except:
                # NOTE: if an exception occurred and we just created the db, remove it again
                import gc
                
                for obj in gc.get_objects():
                    try:
                        # Check if the object is a SQLite connection
                        if isinstance(obj, sql.Connection):
                            try:
                                *_, db_path = obj.cursor().execute("PRAGMA database_list;").fetchone()
                                print(db_path)

                                if db_path == self.db_location:
                                    obj.close()
                                    print(f"Closed connection to {db_path}")
                            except:
                                # NOTE: already closed
                                pass
                    except Exception as e:
                        raise Exception("Error tijdens het sluiten van de database connectie")

                # TODO: still sometimes claims the file is open somewehere
                os.remove(self.db_location)
                raise
        
        self.load_main_tab()
        self.load_other_tabs()

        # Manually recalculate once
        self.reload_tabs()

        self.set_create_button_status()

        with sql.connect(self.db_location) as conn:
            result = conn.execute(f'''
                SELECT table_name, uploaded, edepot_id FROM tables;
            ''').fetchall()

        uploaded_tables = []

        for table_name, uploaded, edepot_id in result:
            if edepot_id is not None:
                self.edepot_ids.append(edepot_id)

            if uploaded:
                uploaded_tables.append(table_name)

        if len(uploaded_tables) > 0:
            if len(uploaded_tables) > len(self.edepot_ids):
                t = threading.Thread(
                    target=self.update_status,
                    args=uploaded_tables
                )
                t.start()

                Dialog(
                    title="Zoeken E-depot",
                    text="Sommige items waren al geupload maar nog niet teruggevonden in het E-depot, deze worden nu verder gezocht.\n\nWanneer de link naar het E-depot beschikbaar is, zal de knop hiervoor actief worden."
                ).exec()
            else:
                self.edepot_available_changed.emit(True)

        self._loaded = True

    def show(self):
        """Show the widget and optionally load the data"""
        if not self._loaded:
            with sql.connect(self.db_location) as conn:
                number_of_rows = conn.execute(f'''SELECT COUNT(*) FROM Overdrachtslijst;''').fetchone()[0]
            if number_of_rows > 1000:
                WarningDialog(
                    title="Lange laadtijd",
                    text=f"Het inladen van {number_of_rows:,} rijen kan een aantal minuten duren. Tijdens het laden is de app niet responsief. Wacht tot de app weer responsief wordt."
                ).exec()
            self.load_items()
            
        return super().show()

    def create_db(self) -> bool:
        os.makedirs(self.storage_base, exist_ok=True)

        if os.path.exists(self.db_location):
            with sql.connect(self.db_location) as conn:
                # ALTERS
                columns = conn.execute("PRAGMA table_info(tables);").fetchall()
                column_names = [column_name for _, column_name, *_ in columns]

                if 'uploaded' not in column_names:
                    conn.execute("""
                        ALTER TABLE tables
                        ADD COLUMN uploaded BOOLEAN;
                    """)

                    conn.execute("""
                        UPDATE tables
                        SET uploaded = CASE
                            WHEN edepot_id IS NOT NULL THEN 1
                            ELSE 0
                        END;
                    """)
            
                if 'URI Serieregister' not in column_names:
                    if 'uri_serieregister' in column_names:
                        conn.execute('ALTER TABLE tables RENAME COLUMN uri_serieregister TO "URI Serieregister";')
                    else:
                        conn.execute('ALTER TABLE tables ADD COLUMN "URI Serieregister" TEXT;')
            
            return False

        with sql.connect(self.db_location) as conn:
            conn.execute("""
            CREATE TABLE tables (
                id INTEGER PRIMARY KEY,
                table_name TEXT,
                "URI Serieregister" TEXT,
                edepot_id TEXT,
                uploaded BOOLEAN DEFAULT 0,

                UNIQUE(table_name)
            );""")

            conn.execute(f"""
            INSERT INTO tables (table_name)
            VALUES ('{self.main_tab}');
            """)

            conn.commit()

        return True

    def load_overdrachtslijst(self):
        wb = load_workbook(
            self.path,
            read_only=True,
            data_only=True,
            keep_links=False,
            rich_text=False,
        )

        if not self.main_tab in wb.sheetnames:
            raise ValueError(f"{self.main_tab} tab missing")

        ws = wb[self.main_tab]
        data = ws.values

        try:
            # TODO: check all required columns?
            while "Doosnr" not in (headers := next(data)):
                pass
        except StopIteration:
            wb.close()
            raise Exception("Geen hoofdingen gevonden in de overdrachtslijst")

        set_headers = set(headers)
        expected_headers = (
            "Beschrijving",
            "Begindatum",
            "Einddatum",
            "Doosnr",
            "URI Serieregister",
        )

        headers = [h.strip() for h in headers if h is not None and h.strip() != ""]
        for h in expected_headers:
            if h not in headers:
                wb.close()
                raise Exception(f"Verwachtte om de kolom '{h}' tegen te komen, maar is niet gevonden.")

        # NOTE: if the headers are part of the location columns, we will allow duplication by giving them unique names
        location_cols = ("Origineel Doosnummer", "Legacy locatie ID", "Legacy range", "Verpakkingstype")
        duplicate_num = 0  # Number to be added to the column name
        new_headers = []
        i = 0

        while i < len(headers):
            header = headers[i]

            if header not in location_cols:
                # NOTE: not a location column
                new_headers.append(header)
                i += 1
                continue
        
            # NOTE: they need to be in order
            for position in range(4):
                if headers[i+position] != location_cols[position]:
                    # Not in order
                    WarningDialog(
                        title="Onduidelijke locatiekolommen",
                        text=f"Een onduidelijke locatiekolom was gevonden '{header}' op positie {i+position+1}.\nLocatie kolommen komen altijd in groepen van 4, in een vaste volgorde (Origineel Doosnummer, Legacy locatie ID, Legacy range, Verpakkingstype).\nDit komt niet overeen met wat er gevonden werd in de overdrachtslijst."
                    ).exec()
                    wb.close()
                    raise Exception("Onduidelijke locatiekolommen")
                
                # In order so far
                new_headers.append(f"{headers[i+position]}{f'_{duplicate_num}' if duplicate_num > 0 else ''}")
            
            duplicate_num += 1
            i += 4

        duplicate_headers = []

        for h in set_headers:
            if new_headers.count(h) > 1:
                duplicate_headers.append(h)

        if len(duplicate_headers) > 0:
            WarningDialog(
                title="Duplicate hoofdingen niet toegestaan",
                text=f"Duplicate hoofdingen zijn niet toegestaan.\n\nDuplicaten gevonden:\n{'\n'.join((f'- {h}' for h in duplicate_headers))}"
            ).exec()
            wb.close()
            raise Exception("Duplicate hoofdingen gevonden")

        # Filter out empty rows
        df = pd.DataFrame(
            (
                r for r in 
                (r[:len(new_headers)] for r in list(data))
                if not all(not bool(v) for v in r)
            ),
            columns=new_headers,
        ).fillna("").astype(str).convert_dtypes()
        wb.close()

        if len(df) > 1000:
            Dialog(
                title="Grote overdrachtslijst",
                text="Je probeert een grote overdrachtslijst te openen, dit kan een aantal minuten duren."
            ).exec()

        for uri, count in df["URI Serieregister"].value_counts().items():
            if uri.strip() != "" and count > 9998:
                raise ValueError(f"Te veel lijnen met 'URI Serieregister' gelijk aan '{uri}'.\nMaximum: 9998\nGevonden: {count}\n\nPas de overdrachtslijst aan alvorens verder te gaan.")

        df["id"] = range(df.shape[0])
        df["series_name"] = ""
        # df.iloc[:50, df.columns.get_loc("series_name")] = "None"

        # NOTE: since Excel deals with datetimes awkwardly, make sure we only have the date part as a string here
        # Only take the date part, if none found, keep original value
        df["Begindatum"] = df["Begindatum"].str.extract(r"(\d{4}-\d{2}-\d{2})", expand=False).fillna(df["Begindatum"])
        df["Einddatum"] = df["Einddatum"].str.extract(r"(\d{4}-\d{2}-\d{2})", expand=False).fillna(df["Einddatum"])

        # NOTE: reorder headers
        cols = df.columns.tolist()
        cols = ["id", "series_name", *(c for c in cols if c not in ("id", "series_name", "URI Serieregister")), "URI Serieregister"]
        df = df[cols]

        con = sql.connect(self.db_location)

        try:
            df.to_sql(
                name=self.main_tab,
                con=con,
                index=False,
                method="multi",
                chunksize=100,
            )
        finally:
            con.close()

    def load_main_tab(self):
        def _set_combobox_items(combobox: QtWidgets.QComboBox) -> None:
            combobox.clear()
            
            # Sort series alphabetically on name, for easier usage
            self.series.sort(key = lambda x:x.name)
            for series in self.series:
                if series.status != "Published":
                    continue

                series_combobox.addItem(series.get_name(), userData=series._id)

        def _set_uris(model: SQLliteModel) -> None:
            uri_pre = self.state.configuration.active_environment.get_serie_register_uri()

            model.all_series_uris = [f"{uri_pre}/{s._id}" for s in self.series]

        container = QtWidgets.QWidget()
        layout = QtWidgets.QGridLayout()
        container.setLayout(layout)
        
        series_combobox = QtWidgets.QComboBox()
        series_combobox.setEditable(True)
        series_combobox.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
        series_combobox.completer().setCompletionMode(
            QtWidgets.QCompleter.PopupCompletion
        )
        series_combobox.completer().setFilterMode(
            QtCore.Qt.MatchFlag.MatchContains
        )
        series_combobox.setMaximumWidth(900)
        _set_combobox_items(series_combobox)

        btn = QtWidgets.QPushButton(text="Voeg toe")
        btn.clicked.connect(
            lambda:
            self.add_to_new(
                name = series_combobox.currentText(),
                series_id = series_combobox.currentData(),
                recalculate=True
            )
        )

        uri_pre = self.state.configuration.active_environment.get_serie_register_uri()
        model = SQLliteModel(
            self.main_tab,
            state=self.state,
            db_name=self.db_location,
            is_main=True,
        )
        _set_uris(model)

        proxy_model = CustomSortFilterModel()
        proxy_model.setSourceModel(model)
        model.bad_rows_changed.connect(self.set_create_button_status)
        self.main_table.setModel(proxy_model)
        
        self.unassigned_only_checkbox = QtWidgets.QCheckBox(text="Toon enkel rijen zonder serie")
        self.unassigned_only_checkbox.stateChanged.connect(self._filter_unassigned)
        
        default_sorting_button = QtWidgets.QPushButton(text="Reset sortering")
        default_sorting_button.clicked.connect(proxy_model.reset_sorting)

        layout.addWidget(btn, 0, 0)
        layout.addWidget(series_combobox, 0, 1, 1, 3)
        layout.addWidget(self.unassigned_only_checkbox, 1, 0)
        layout.addWidget(default_sorting_button, 1, 4)
        layout.addWidget(self.main_table, 2, 0, 1, 5)

        self.tab_widget.addTab(container, self.main_tab)
        self.tabs[self.main_tab] = {
            "container": container,
            "table": self.main_table,
            "model": model,
            "proxy": proxy_model,
        }

        # NOTE: map all the URI Serieregisters
        with sql.connect(self.db_location) as conn:
            unique_serie_uris = [
                uri for uri, *_ in 
                conn.execute(f"""SELECT "URI Serieregister" FROM {self.main_tab} GROUP BY "URI Serieregister";""").fetchall()
            ]

            uri_index_maps = {}

            for serie_uri in unique_serie_uris:
                uri_index_maps[serie_uri] = [
                    i for i, *_ in
                    conn.execute(f"""SELECT id FROM {self.main_tab} WHERE "URI Serieregister" = '{serie_uri}';""").fetchall()
                ]

        # NOTE: set all the series_names where the series_id matches one we got
        for uri, indexes in uri_index_maps.items():
            match = [s for s in self.series if s._id == uri.split(uri_pre + "/")[-1]]

            if len(match) != 1:
                continue

            series = match[0]

            self.add_to_new(name=series.get_name(), series_id=series._id, mapping_ids=indexes, recalculate=False)

    def load_other_tabs(self):
        with sql.connect(self.db_location) as conn:
            tables = conn.execute(f"""SELECT table_name, "URI Serieregister" FROM tables WHERE table_name != '{self.main_tab}';""")
        
        # NOTE: set all the series_names where the series_id matches one we got
        uri_pre = self.state.configuration.active_environment.get_serie_register_uri()

        for table_name, uri_serieregister in tables:
            match = [s for s in self.series if s._id == uri_serieregister.split(uri_pre + "/")[-1]]

            if len(match) != 1:
                continue

            series = match[0]

            # NOTE: remove leading and trailing quotes
            self.create_tab(name=table_name[1:-1], series_id=series._id)

    def add_to_new(self, name: str, series_id: str, mapping_ids: list[int] = None, recalculate=True):
        # NOTE: only thing not allowed is quotes
        name = name.strip().replace('"', "").replace("'", "")

        # No funny business
        if name == "" or name == self.main_tab:
            return

        amount_of_rows_to_add = 0

        if mapping_ids is None:
            proxy: CustomSortFilterModel = self.tabs[self.main_tab]["proxy"]

            # NOTE: filter out hidden rows here
            selected_rows = [
                str(proxy.mapToSource(r).row())
                for r in self.main_table.selectionModel().selectedRows()
            ]

            if len(selected_rows) == 0:
                return

            selected_rows_str = ", ".join(selected_rows)

            amount_of_rows_to_add = len(selected_rows)

            if amount_of_rows_to_add > 9998:
                raise ValueError(f"Te veel lijnen worden in de serie toegevoegd.\nMaximum: 9998\nGevonden: {amount_of_rows_to_add}")
            elif amount_of_rows_to_add > 1000:
                Dialog(
                    title="Veel rijen aanpassen",
                    text="Je bent veel rijen tegelijk aan het aanpassen, dit kan een aantal minuten duren."
                ).exec()
        else:
            selected_rows_str = ", ".join(str(i) for i in mapping_ids)

        uri = f"{self.state.configuration.active_environment.get_serie_register_uri()}/{series_id}"
        tab_exists = False

        with sql.connect(self.db_location) as conn:
            # Get main table columns for later
            result = conn.execute(f'pragma table_info({self.main_tab});').fetchall()
            main_table_columns = [col for _id, col, _type, *_ in result]

            # Check if table exists
            result = conn.execute(f'pragma table_info("{name}");').fetchall()

            # Create table
            if not result:
                import_sjabloon = APIController.get_import_template(self.state.configuration, series_id=series_id)

                columns = pd.read_excel(import_sjabloon, dtype=str, engine="openpyxl").columns.to_list()

                # NOTE: if the main table contains duplicate columns of existing ones, we want to take those along
                is_duplicate_column = lambda c: len(c.split("_", 1)) == 2 and c.split("_", 1)[0] in columns
                duplicate_main_columns_mapping = {
                    c: c.split("_")[0]
                    for c in main_table_columns if is_duplicate_column(c)
                }

                # NOTE: add the duplicate columns in the right spot
                new_columns = []

                for column in columns:
                    new_columns.append(column)

                    # NOTE: Since the order of these is a bit more restricted, we need to have more logic for it
                    if column in ("Origineel Doosnummer", "Legacy locatie ID", "Legacy range"):
                        continue
                    elif column == "Verpakkingstype":
                        # NOTE: now add all the duplicate location-columns in order
                        duplicates = {
                            "Origineel Doosnummer": [],
                            "Legacy locatie ID": [],
                            "Legacy range": [],
                            "Verpakkingstype": [],
                        }

                        for duplicate_location_column, original_column in duplicate_main_columns_mapping.items():
                            if original_column in duplicates:
                                duplicates[original_column].append(duplicate_location_column)

                        # NOTE: in case they were not sorted properly in the metadata, we do it now
                        duplicates["Origineel Doosnummer"] = sorted(duplicates["Origineel Doosnummer"], key=natural_keys)
                        duplicates["Legacy locatie ID"] = sorted(duplicates["Legacy locatie ID"], key=natural_keys)
                        duplicates["Legacy range"] = sorted(duplicates["Legacy range"], key=natural_keys)
                        duplicates["Verpakkingstype"] = sorted(duplicates["Verpakkingstype"], key=natural_keys)

                        # NOTE: zip will automatically take the shortest list
                        for loc_col_1, loc_col_2, loc_col_3, loc_col_4 in zip(*duplicates.values()):
                            new_columns.append(loc_col_1)
                            new_columns.append(loc_col_2)
                            new_columns.append(loc_col_3)
                            new_columns.append(loc_col_4)

                        continue

                    for duplicate_column in [dc for dc, c in duplicate_main_columns_mapping.items() if c == column]:
                        new_columns.append(duplicate_column)

                columns = new_columns

                conn.execute(f"""
                CREATE TABLE IF NOT EXISTS "{name}" (
                    id INTEGER PRIMARY KEY,

                    main_id INTEGER NOT NULL,

                    {",\n\t".join(f'"{c}" TEXT' for c in columns)}
                );""")

                # Update the tables table
                conn.execute(f"""
                    INSERT OR IGNORE INTO tables (table_name, "URI Serieregister")
                    VALUES ('"{name}"', '{uri}');
                """)
            else:
                tab_exists = True
                columns = [col for _id, col, _type, *_ in result]

            # Check if we don't have to many rows already
            result, *_ = conn.execute(f'SELECT count() FROM "{name}";').fetchone()
            
            if result + amount_of_rows_to_add > 9998:
                raise ValueError(f"Te veel lijnen worden in de serie toegevoegd.\nMaximum: 9998\nGevonden: {result + amount_of_rows_to_add}")

            # Remove where needed
            cursor = conn.execute(f"""
                SELECT id, series_name
                FROM {self.main_tab}
                WHERE id IN ({selected_rows_str})
                  AND series_name != '"{name}"'
                  AND series_name != '';
            """)

            for main_id, table in cursor.fetchall():
                # NOTE: table already contains ""-marks
                tab = table[1:-1]

                # NOTE: table could be non-existent if series_name was somehow filled in accidentally
                try:
                    conn.execute(f"""
                        DELETE FROM {table}
                        WHERE main_id={main_id};
                    """)
                    
                    rows = conn.execute(f"""
                        SELECT count() FROM {table};
                    """).fetchone()[0]

                    if rows == 0:
                        conn.execute(f"""
                            DROP TABLE {table};
                        """)
                        
                        conn.execute(f"""
                            DELETE FROM tables
                            WHERE table_name='{table}';
                        """)

                        container: QtWidgets.QWidget = self.tabs[tab]["container"]
                        _table: TableView = self.tabs[tab]["table"]
                        model: SQLliteModel = self.tabs[tab]["model"]
                        model.bad_rows_changed.disconnect()

                        self.tab_widget.removeTab(self.tab_widget.indexOf(container))
                        model.deleteLater()
                        _table.deleteLater()
                        container.deleteLater()

                        self.tabs.pop(tab)

                        continue

                    # Recalculate shape for table
                    model: SQLliteModel = self.tabs[tab]["model"]
                    model.row_count = rows
                except:
                    pass

            # Insert where needed
            # NOTE: map names that are a 1-to-1 match
            fixed_mapping = {
                'main_id': 'id',
                # 'Type': 'dossier',
                'Analoog?': '\'ja\'',
                'Path in SIP': '"Beschrijving"',
                'Naam': '"Beschrijving"',
                'Openingsdatum': '"Begindatum"',
                'Sluitingsdatum': '"Einddatum"',
                'Origineel Doosnummer': f'''
                    CASE WHEN "Doosnr" = '' OR "Doosnr" GLOB '[0-9][0-9][0-9][0-9]' OR "Doosnr" GLOB '[0-9][0-9][0-9]' OR "Doosnr" GLOB '[0-9][0-9]' OR "Doosnr" GLOB '[0-9]'
                        THEN substr(\'0000\' || "Doosnr", -4, 4) || \'/{self.overdrachtslijst_name.replace("'", "''")}\'
                        ELSE "Doosnr" || \'/{self.overdrachtslijst_name.replace("'", "''")}\'
                    END
                '''
            }

            matching_columns = set(columns) & set(main_table_columns)

            # NOTE: only map "Beschrijving" if we are already mapping "Naam" manually
            if "Naam" not in matching_columns and "Beschrijving" in matching_columns:
                matching_columns.remove("Beschrijving")

            matching_cols = [f'"{c}"' for c in matching_columns if c != "id"]

            # NOTE: values from dynamic mapping are more important, since it's user input
            fixed_mapping_cols = [f'"{c}"' for c in fixed_mapping if c not in matching_columns]
            fixed_mapping_values = [
                # NOTE: we might need to escape single quotes
                # v.replace("'", "''") if c != "Origineel Doosnummer" else v
                v
                for c, v in fixed_mapping.items()
                if c not in matching_columns
            ]

            mapping_cols = matching_cols + fixed_mapping_cols
            mapping_values = matching_cols + fixed_mapping_values

            cols_str = ", ".join(mapping_cols)
            values_str = ", ".join(mapping_values)

            conn.execute(f"""
                INSERT INTO "{name}" ({cols_str})
                SELECT {values_str}
                FROM {self.main_tab}
                WHERE id in ({selected_rows_str})
                  AND (series_name != '"{name}"' OR series_name IS NULL OR series_name == '');
            """)
            
            # Update the main table to show correct linking
            conn.execute(f"""
                UPDATE {self.main_tab}
                SET series_name='"{name}"',
                    "URI Serieregister"='{uri}'
                WHERE id IN ({selected_rows_str});
            """)

            conn.commit()

        if mapping_ids:
            # NOTE: we added this automatically, don't add the tab here
            pass
        elif not tab_exists:
            self.create_tab(name, series_id)

        if recalculate:
            # Update the graphical side for all tables involved
            self.reload_tabs(new_tab=None if tab_exists else name)

    def create_tab(self, name: str, series_id: str):
        container = QtWidgets.QWidget()
        layout = QtWidgets.QGridLayout()
        container.setLayout(layout)

        series_label = QtWidgets.QLabel(text=name)
        
        duplicate_trefwoord_column_button = QtWidgets.QPushButton(text="Dupliceer trefwoorden_vrij kolom")
        duplicate_location_column_button = QtWidgets.QPushButton(text="Dupliceer locatie kolommen")
        duplicate_location_column_button.setHidden(self.state.configuration.active_role == "klant")
        self.configuration_changed.connect(lambda: self.hide_or_show_button(duplicate_location_column_button))

        table_view = TableView()
        bad_rows_checkbox = QtWidgets.QCheckBox(text="Toon enkel rijen met fouten")
        default_sorting_button = QtWidgets.QPushButton(text="Reset sortering")

        layout.addWidget(series_label, 0, 0, 1, 2)
        layout.addWidget(duplicate_trefwoord_column_button, 1, 0)
        layout.addWidget(duplicate_location_column_button, 1, 1)
        layout.addWidget(bad_rows_checkbox, 2, 0)
        layout.addWidget(default_sorting_button, 2, 4)
        layout.addWidget(table_view, 3, 0, 1, 5)

        model = SQLliteModel(
            name,
            state=self.state,
            db_name=self.db_location,
            series_id=series_id
        )
        proxy_model = CustomSortFilterModel()
        proxy_model.setSourceModel(model)
        table_view.setModel(proxy_model)
        
        default_sorting_button.clicked.connect(proxy_model.reset_sorting)

        load_bestandscontrole_button = QtWidgets.QPushButton(text="Laad bestandscontrole lijst")
        load_bestandscontrole_button.clicked.connect(lambda: self.load_bestandscontrole(model=model))
        load_bestandscontrole_button.setHidden(self.state.configuration.active_role == "klant")
        self.configuration_changed.connect(lambda: self.hide_or_show_button(load_bestandscontrole_button))
        layout.addWidget(load_bestandscontrole_button, 1, 4)

        bad_rows_checkbox.stateChanged.connect(lambda checkstate: self._filter_bad_rows(checkstate, self.tabs[name]["table"]))
        model.bad_rows_changed.connect(lambda: self._filter_bad_rows(bad_rows_checkbox.checkState().value, table_view))
        model.layoutChanged.connect(lambda: self._filter_bad_rows(bad_rows_checkbox.checkState().value, table_view))
        model.modelReset.connect(lambda: self._filter_bad_rows(bad_rows_checkbox.checkState().value, table_view))
        model.bad_rows_changed.connect(self.set_create_button_status)
        self.configuration_changed.connect(lambda: self._filter_bad_rows(bad_rows_checkbox.checkState().value, table_view))

        self.tab_widget.addTab(container, name)
        self.tabs[name] = {
            "container": container,
            "table": table_view,
            "model": model,
            "proxy": proxy_model,
        }

        duplicate_trefwoord_column_button.clicked.connect(lambda: self.add_column(name))
        duplicate_location_column_button.clicked.connect(lambda: self.add_column(name, location_cols=True))

    def closeEvent(self, event):
        models: list[SQLliteModel] = [t["model"] for t in self.tabs.values()]

        # NOTE: only ask the user if data has actually changed
        if not any(m.has_changed for m in models):
            event.accept()
            return

        dialog = YesNoDialog(
            title="Overdrachtslijst sluiten",
            text="Wil je de huidige data opslaan?\nZonder opslaan gaat de nieuwe data verloren bij heropstarten van de applicatie."
        )
        dialog.exec()

        if dialog.result():
            self.save_tabs()

        event.accept()

    def save_tabs(self) -> None:
        for reference in self.tabs.values():
            table_view = reference["table"]

            if table_view == self.main_table:
                continue

            model: SQLliteModel = reference["model"]

            model.save_data()
            model.has_changed = False

        self.close()

    def reload_tabs(self, new_tab: str=None) -> None:
        # A wait message on the Opslaan button
        self.save_button.setText("Wachten op inladen van data...")
        self.save_button.setStyleSheet("background-color: lightblue;")
        
        # Dict to store all threads
        self.data_loading_threads = {}
        for table_name, reference in self.tabs.items():
            # Reload all the data
            model: SQLliteModel = reference["model"]

            # Start background thread
            thread = QThread()
            worker = DataLoadingWorker(table_name, model)
            worker.moveToThread(thread)

            thread.started.connect(worker.run)
            worker.finished.connect(self.on_load_finished)
            worker.finished.connect(worker.deleteLater)
            thread.finished.connect(thread.deleteLater)

            self.data_loading_threads[table_name] = {"thread":thread, "worker":worker, "finished":False}
            thread.start()

            # Only do this if the tab is new, or we are reloading all tabs
            if new_tab is None or new_tab == table_name:
                tab_view: TableView = reference["table"]

                with sql.connect(self.db_location) as conn:
                    # NOTE: figure out which columns to hide (could be multiple due to duplications)
                    cursor = conn.execute(f"pragma table_info(\"{table_name}\");")

                    columns = cursor.fetchall()

                # Show every column
                for i in range(len(columns)):
                    tab_view.showColumn(i)
                    model.mark_column_as_hidden(i, hidden=False)

                # Hide id and main_id columns where applicable
                for i, column_name, *_ in columns:
                    if column_name in ("id", "main_id"):
                        tab_view.hideColumn(i)
                        model.mark_column_as_hidden(i)

                if self.state.configuration.active_role == "klant":
                    cols_to_skip = ("Origineel Doosnummer", "Legacy locatie ID", "Legacy range", "Verpakkingstype")

                    for i, column_name, *_ in columns:
                        if any(c in column_name for c in cols_to_skip):
                            tab_view.hideColumn(i)
                            model.mark_column_as_hidden(i)

        self._filter_unassigned(self.unassigned_only_checkbox.checkState().value)

    @Slot(str)
    def on_load_finished(self, table_name):
        """Run when loading of the data for all tables on the tab are finished"""
        self.data_loading_threads[table_name]["thread"].quit()
        self.data_loading_threads[table_name]["finished"] = True

        if all([v["finished"] for v in self.data_loading_threads.values()]):
            # Remove wait message
            self.save_button.setText("Opslaan")
            self.save_button.setStyleSheet("")

    def set_create_button_status(self, *_) -> None:
        for reference in self.tabs.values():
            model: SQLliteModel = reference["model"]

            red_colors = [_ for c in model.colors.values() if c == CellColor.RED]

            if len(red_colors) > 0:
                self.create_sips_button.setEnabled(False)
                self.can_upload = False
                self.can_upload_changed.emit(False)
                return

        self.create_sips_button.setEnabled(True)
        self.can_upload = True
        self.can_upload_changed.emit(True)

    def hide_or_show_button(self, button: QtWidgets.QPushButton) -> None:
        button.setHidden(self.state.configuration.active_role == "klant")

    def add_column(self, table: str, location_cols: bool=False) -> None:
        def find_next(old_columns: list, name: str) -> tuple[str, str]:
            # Returns new column name, and the previous one to look for
            value = 1

            while (new_name := f'{name}_{value}') in old_columns:
                value += 1

            return new_name, f'{name}_{value-1}' if value > 1 else name

        with sql.connect(self.db_location) as conn:
            # NOTE: since we need to reset the whole table definition
            # Rename the table
            conn.execute(f'ALTER TABLE "{table}" RENAME TO "_old_{table}";')

            # Get old table definition
            cursor = conn.execute(f'pragma table_info("_old_{table}");')

            old_definitions = [(name, _type) for _, name, _type, *_ in cursor.fetchall()]
            old_columns = [d[0] for d in old_definitions]

            # Find where to input the new column
            new_column_name, previous_column_name = find_next(old_columns, "Origineel Doosnummer" if location_cols else "Trefwoorden_vrij")
            
            if location_cols:
                # We need to make sure we select the last one in the set
                previous_column_name = previous_column_name.replace("Origineel Doosnummer", "Verpakkingstype")

            # Create the new definitions
            new_definitions = []

            for column_name, column_type in old_definitions:
                new_definitions.append(f'"{column_name}" {column_type}')

                if column_name == previous_column_name:
                    if location_cols:
                        columns = ("Origineel Doosnummer", "Legacy locatie ID", "Legacy range", "Verpakkingstype")
                        number = new_column_name.split('_')[-1]

                        for c in columns:
                            new_definitions.append(f'"{c}_{number}" {column_type}')
                    else:
                        new_definitions.append(f'"{new_column_name}" {column_type}')

            # Create new table
            conn.execute(f'''
                CREATE TABLE "{table}" (
                    {",\n\t".join(new_definitions)}
                );
            ''')

            # Fill table with old values
            conn.execute(f'''
                INSERT INTO "{table}" ({", ".join(f'"{c}"' for c in old_columns)})
                SELECT {", ".join(f'"{c}"' for c in old_columns)}
                FROM "_old_{table}";
            ''')

            conn.execute(f'DROP TABLE "_old_{table}";')

        model: SQLliteModel = self.tabs[table]["model"]
        model.get_data()

    def _filter_unassigned(self, state: QtCore.Qt.CheckState) -> None:
        model: CustomSortFilterModel = self.main_table.model(proxy=True)
        
        if state == QtCore.Qt.CheckState.Checked.value:
            model.add_filter(CustomSortFilterModel.ROWS_WITHOUT_SERIES_FILTER)
        else:
            model.remove_filter(CustomSortFilterModel.ROWS_WITHOUT_SERIES_FILTER)

    def _filter_bad_rows(self, state: QtCore.Qt.CheckState, table_view: TableView) -> None:
        model: CustomSortFilterModel = table_view.model(proxy=True)

        if state == QtCore.Qt.CheckState.Checked.value:
            model.add_filter(CustomSortFilterModel.BAD_ROWS_FILTER)
        else:
            model.remove_filter(CustomSortFilterModel.BAD_ROWS_FILTER)

    def create_sips(self) -> None:
        def _col_index_to_xslx_col(col_index: int) -> str:
            # NOTE: this only supports up to ZZ for now
            first_letter_value = (col_index // 26) - 1

            if first_letter_value >= 26:
                raise ValueError("There are too many columns")
            if first_letter_value == -1:
                return chr(65 + col_index)

            first_letter = chr(65 + first_letter_value)
            second_letter = chr(65 + col_index % 26)
            
            return f"{first_letter}{second_letter}"

        storage_location = self.state.configuration.misc.save_location
        sjabloon_base_path = os.path.join(storage_location, FileController.IMPORT_TEMPLATE_STORAGE)
        sip_storage_path = os.path.join(storage_location, FileController.SIP_STORAGE)
        grid_storage_path = os.path.join(storage_location, FileController.GRID_STORAGE)

        os.makedirs(sip_storage_path, exist_ok=True)
        os.makedirs(grid_storage_path, exist_ok=True)
        os.makedirs(sjabloon_base_path, exist_ok=True)

        for series_name, reference in self.tabs.items():
            if series_name == self.main_tab:
                continue

            model: SQLliteModel = reference["model"]
            model.save_data()

            # Download the import_template
            import_template_loc = APIController.get_import_template(self.state.configuration, series_id=model.series_id)

            # Copy import_template to grid_storage
            temp_loc = os.path.join(grid_storage_path, f"temp_{model.series_id}.xlsx")

            wb = load_workbook(import_template_loc)
            ws = wb["Details"]

            data = model.raw_data

            # NOTE: make sure we don't exceed 9999 lines (counting the heading as one)
            if len(data) > 9999 - 1:
                WarningDialog(
                    title="Te veel lijnen",
                    text=f"Er zijn te veel lijnen in de tab '{series_name}'.\n\nMaximum: 9999\nGevonden: {len(data) + 1}"
                ).exec()
                return

            # NOTE: overwrite the headers
            for col_index, col in model.columns.items():
                # NOTE: skip id and main_id
                if col_index < 2:
                    continue

                # NOTE: duplicate columns generate like <col>_<n>
                # duplicate columns read generate like <col>.<n>
                re_match = re.match(r"(.*)[_.]\d+$", col)

                if re_match:
                    col = re_match.group(1)

                ws[f"{_col_index_to_xslx_col(col_index - 2)}1"] = col

            for row_index, row in enumerate(data):
                for col_index, value in enumerate(row):
                    # NOTE: skip id and main_id
                    if col_index < 2:
                        continue

                    ws[f"{_col_index_to_xslx_col(col_index - 2)}{row_index+2}"] = value

            wb.save(temp_loc)
            wb.close()

            ol_name = self.overdrachtslijst_name[:185]

            sip_location = os.path.join(sip_storage_path, f"{model.series_id}-{ol_name}-SIPC.zip")
            md5_location = os.path.join(sip_storage_path, f"{model.series_id}-{ol_name}-SIPC.xml")

            with zipfile.ZipFile(
                sip_location, "w", compression=zipfile.ZIP_DEFLATED
            ) as zfile:
                zfile.write(
                    temp_loc,
                    "Metadata.xlsx"
                )

            md5 = hashlib.md5(open(sip_location, "rb").read()).hexdigest()

            side_car_info = """<?xml version="1.0" encoding="UTF-8"?>
<mhs:Sidecar xmlns:mhs="https://zeticon.mediahaven.com/metadata/20.3/mhs/" version="20.3" xmlns:mh="https://zeticon.mediahaven.com/metadata/20.3/mh/">
    <mhs:Technical>
            <mh:Md5>{md5}</mh:Md5>
    </mhs:Technical>
</mhs:Sidecar>""".format(
                md5=md5
            )

            with open(md5_location, "w", encoding="utf-8") as f:
                f.write(side_car_info)

            os.remove(temp_loc)
        
            # NOTE; set the edepot_id to empty again, since it hasn't been uploaded in it's current form
            with sql.connect(self.db_location) as conn:
                conn.execute(f'''
                    UPDATE tables
                    SET edepot_id=null,
                        uploaded=0
                    WHERE table_name='"{series_name}"';
                ''')

        self.can_upload = True
        self.can_upload_changed.emit(True)
        self.edepot_available_changed.emit(False)

        Dialog(
            title="SIPs aangemaakt",
            text="SIPS zijn aangemaakt voor de overdrachtslijst."
        ).exec()

        self.close()

    def upload_sips(self) -> None:
        # NOTE: ask which ones to upload
        tabs: list[tuple[str, TableView]] = []
        
        for tab, reference in self.tabs.items():
            if tab == self.main_tab:
                continue

            tabs.append((tab, reference["model"]))

        dialog = ChoiceDialog(
            title="Uploaden",
            text="Kies voor welke series je een SIP wilt uploaden",
            choices=[t for t, _ in tabs],
            default_selected=True,
        )
        dialog.exec()

        if dialog.result():
            tabs_to_upload = dialog.get_selected_choices()
            tabs_to_upload = [(tab, model) for tab, model in tabs if tab in tabs_to_upload]

            if len(tabs_to_upload) == 0:
                return
        else:
            return

        env = self.state.configuration.active_environment
        if not env.has_ftps_credentials():
            WarningDialog(
                title="Connectie fout",
                text=f"Je FTPS connectie gegevens staan niet in orde voor omgeving '{self.sip.environment.name}'",
            ).exec()
            return

        storage_location = self.state.configuration.misc.save_location
        sip_storage_path = os.path.join(storage_location, FileController.SIP_STORAGE)

        for series_name, model in tabs_to_upload:
            model: SQLliteModel = model

            ol_name = self.overdrachtslijst_name[:185]

            sip_location = os.path.join(sip_storage_path, f"{model.series_id}-{ol_name}-SIPC.zip")
            md5_location = os.path.join(sip_storage_path, f"{model.series_id}-{ol_name}-SIPC.xml")

            if not os.path.exists(sip_location) or not os.path.exists(md5_location):
                self.create_sips()
            
            try:
                with ftplib.FTP_TLS(
                    env.ftps_url,
                    env.ftps_username,
                    env.ftps_password,
                ) as session:
                    session.prot_p()

                    with open(sip_location, "rb") as f:
                        session.storbinary(f"STOR {model.series_id}-{ol_name}-SIPC.zip", f)
                    with open(md5_location, "rb") as f:
                        session.storbinary(f"STOR {model.series_id}-{ol_name}-SIPC.xml", f)
            except ftplib.error_perm:
                WarningDialog(
                    title="Connectie fout",
                    text=f"Je FTPS connectie login gegevens staan niet in orde voor omgeving '{self.sip.environment.name}'",
                ).exec()
                return
            except socket.gaierror:
                WarningDialog(
                    title="Connectie fout",
                    text=f"Je FTPS connectie url staat niet in orde voor omgeving '{self.sip.environment.name}'",
                ).exec()
                return
 
            with sql.connect(self.db_location) as conn:
                conn.execute(f'''
                    UPDATE tables
                    SET uploaded=1
                    WHERE table_name='"{series_name}"';
                ''')

        t = threading.Thread(
            target=self.update_status,
            args=[f'"{tab}"' for tab, _ in tabs_to_upload]
        )
        t.start()

        Dialog(
            title="Upload geslaagd",
            text="Upload voor de overdrachtslijst is geslaagd.\nDe overdrachtslijst blijft in de lijst staan zolang hij op je computer staat.\nOm hem weg te halen, verwijder de correcte database uit de bestandslocatie.\n\nWanneer de items op het E-depot staan zal de knop hiervoor beschikbaar worden.\n\nZodra de link naar het E-depot beschikbaar is, zal de knop ook actief worden."
        ).exec()
        self.can_upload_changed.emit(False)

    def update_status(self, *tabs: Iterable[str]) -> None:
        for series_name, reference in self.tabs.items():
            if series_name == self.main_tab:
                continue

            if f'"{series_name}"' not in tabs:
                continue

            model: SQLliteModel = reference["model"]
            edepot_id = None

            while edepot_id is None:
                print(f"Starting to check for {self.overdrachtslijst_name} - {series_name}")
                # NOTE: wait some time for the edepot to pick them up
                time.sleep(10)

                edepot_id = APIController.get_sip_id_for_name(
                    self.state.configuration.active_environment,
                    f"{model.series_id}-{self.overdrachtslijst_name}-SIPC.zip"
                )

            print(f"id found: {edepot_id} for {model.series_id}-{self.overdrachtslijst_name}")

            with sql.connect(self.db_location) as conn:
                conn.execute(f'''
                    UPDATE tables
                    SET edepot_id='{edepot_id}'
                    WHERE table_name='"{series_name}"';
                ''')

                if edepot_id not in self.edepot_ids:
                    self.edepot_ids.append(edepot_id)

        self.edepot_available_changed.emit(True)

    def load_bestandscontrole(self, model: SQLliteModel) -> None:
        controller = self.application.bestands_controle_lijst_controller

        if not controller.valid:
            WarningDialog(
                title="Bestandscontrole lijst is niet geldig",
                text="De bestandscontrole lijst is niet geldig, bekijk of het pad juist staat, en/of het bestand in orde is."
            ).exec()
            return

        values = controller.get_values(overdrachtslijst_name=self.overdrachtslijst_name)

        if values is None:
            # NOTE: a warning has already been shown to the user
            return

        if model.row_count > 1000:
            WarningDialog(
                title="Trage actie",
                text="Omdat er veel rijen moeten veranderen kan de actie tot enkele minuten duren."
            ).exec()

        col_indeces = []

        for i, col in model.columns.items():
            if col in ("Legacy locatie ID", "Legacy range", "Verpakkingstype"):
                col_indeces.append(i)

                if col == "Legacy locatie ID":
                    new_val = values[controller.list_start_column]
                elif col == "Legacy range":
                    new_val = values[controller.list_end_column]
                elif col == "Verpakkingstype":
                    new_val = values[controller.doos_type_column]
                
                non_empty_val = new_val != ""

                for r in range(0, model.row_count):
                    model.set_value(model.index(r, i), new_value=new_val)

                    if non_empty_val:
                        # NOTE: manually unmark, since this is waaaaay faster
                        model._mark_cell(r, i)

        model.dataChanged.emit(
            model.index(0, min(col_indeces)),
            model.index(model.row_count, max(col_indeces))
        )


class MigrationListView(QtWidgets.QWidget):
    def __init__(self, tab_ui: TabUI):
        super().__init__()

        self.application: Application = QtWidgets.QApplication.instance()
        self.state: State = self.application.state

        self.overdrachtslijst_name = tab_ui.overdrachtslijst_name
        self.tab_ui = tab_ui

        layout = QtWidgets.QGridLayout()
        self.setLayout(layout)

        title = QtWidgets.QLabel(text=self.overdrachtslijst_name)

        open_button = QtWidgets.QPushButton(text="Open")
        open_button.clicked.connect(self.tab_ui.show)

        self.upload_button = QtWidgets.QPushButton(text="Upload")
        self.upload_button.clicked.connect(self.tab_ui.upload_sips)
        self.upload_button.setHidden(self.state.configuration.active_role == "klant")
        self.tab_ui.configuration_changed.connect(lambda: self.upload_button.setHidden(self.state.configuration.active_role == "klant"))

        self.upload_button.setEnabled(self.tab_ui.can_upload)
        self.tab_ui.can_upload_changed.connect(self.upload_button.setEnabled)

        self.edepot_button = QtWidgets.QPushButton(text="Open E-depot")
        self.edepot_button.clicked.connect(self.open_edepot_clicked)
        self.edepot_button.setHidden(self.state.configuration.active_role == "klant")
        self.tab_ui.configuration_changed.connect(lambda: self.edepot_button.setHidden(self.state.configuration.active_role == "klant"))

        self.edepot_button.setEnabled(len(self.tab_ui.edepot_ids) > 0)

        self.tab_ui.edepot_available_changed.connect(self.edepot_button.setEnabled)

        layout.addWidget(title, 0, 0, 1, 3)
        layout.addWidget(open_button, 0, 3)
        layout.addWidget(self.upload_button, 1, 3)
        layout.addWidget(self.edepot_button, 2, 3)

    def open_edepot_clicked(self) -> None:
        for edepot_id in self.tab_ui.edepot_ids:
            if edepot_id is not None:
                os.startfile(
                    f"{self.state.configuration.active_environment.api_url}/input/processing-list/{edepot_id}"
                )


class DataLoadingWorker(QObject):
    finished = Signal(str)

    def __init__(self, table_name:str, model:SQLliteModel):
        super().__init__()
        self.model = model
        self.table_name = table_name

    def run(self):
        self.model.get_data()
        self.finished.emit(self.table_name)
