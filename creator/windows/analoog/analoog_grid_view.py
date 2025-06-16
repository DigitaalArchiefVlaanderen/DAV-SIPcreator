# NOTE: quite a lot of code duplication here with windows.grid_view
# Probably a cleaner way of handling both cases at the same time
import os

from PySide6 import QtWidgets, QtCore, QtGui
import sqlite3 as sql

from creator.application import Application

from creator.controllers.api_controller import APIController

from creator.widgets.tableview_widget import TableView
from creator.widgets.toolbar import Toolbar
from creator.widgets.warning_dialog import WarningDialog
from creator.widgets.dialog import YesNoDialog
from creator.widgets.dialog import Dialog

from creator.utils.state import State
from creator.utils.path_loader import resource_path
from creator.utils.analoog.list_item import ListItem
from creator.utils.table.list_table_model import ListTableModel
from creator.utils.proxymodel import CustomSortFilterModel


import os
import re
import threading
import time
from typing import Iterable

import zipfile
import hashlib
import ftplib
import socket
from openpyxl import load_workbook
from PySide6 import QtWidgets, QtCore, QtGui
import pandas as pd
import sqlite3 as sql
from pathlib import Path

from creator.application import Application, natural_keys

from creator.widgets.main_widgets.main_widget import MainWidget
from creator.widgets.searchable_list_widget import SearchableListWidget
from creator.widgets.tableview_widget import TableView
from creator.widgets.toolbar import Toolbar
from creator.widgets.dialog import YesNoDialog, Dialog, ChoiceDialog
from creator.widgets.warning_dialog import WarningDialog

from creator.controllers.api_controller import APIController
from creator.controllers.file_controller import FileController

from creator.utils.state import State
from creator.utils.sqlitemodel import SQLliteModel, CellColor
from creator.utils.proxymodel import CustomSortFilterModel
from creator.utils.path_loader import resource_path

from creator.windows.mainwindow import MainWindow


class AnaloogGridView(QtWidgets.QMainWindow):
    def __init__(self, list_item: ListItem):
        super().__init__()

        self.application: Application = QtWidgets.QApplication.instance()
        self.state: State = self.application.state

        self.list_item = list_item

        self.saved = True

    def _reset_saved(self) -> None:
        self.saved = False

    def setup_ui(self) -> None:
        self.resize(800, 600)
        self.setWindowTitle(self.list_item.name)

        self.toolbar = Toolbar()
        self.addToolBar(self.toolbar)

        self.setWindowIcon(QtGui.QIcon(resource_path("logo.ico")))

        central_widget = QtWidgets.QWidget()
        self.setCentralWidget(central_widget)

        grid_layout = QtWidgets.QGridLayout()
        central_widget.setLayout(grid_layout)

        series_label = QtWidgets.QLabel(text=self.list_item.grid.series.get_name())
        self.default_sorting_button = QtWidgets.QPushButton(text="Reset sortering")

        self.show_bad_rows_checkbox = QtWidgets.QCheckBox(
            text="Toon enkel rijen met fouten"
        )
        self.show_bad_rows_checkbox.stateChanged.connect(self._bad_rows_clicked)

        self.table_view = TableView(is_analoog=True)

        self.model = ListTableModel(list_item=self.list_item)
        self.model.first_open()
        self.model.dataChanged.connect(lambda: self._reset_saved())
        self.proxy_model = CustomSortFilterModel()
        self.proxy_model.setSourceModel(self.model)
        self.table_view.setModel(self.proxy_model)

        self.default_sorting_button.clicked.connect(lambda: self.proxy_model.reset_sorting())

        id_col = [i for i, c in self.model.columns.items() if c == "_id"][0]
        self.table_view.hideColumn(id_col)

        save_button = QtWidgets.QPushButton(text="Opslaan")
        save_button.clicked.connect(self.save_button_click)

        self.create_sip_button = QtWidgets.QPushButton(text="Maak SIP")
        self.create_sip_button.clicked.connect(self.create_sip_click)
        self.create_sip_button.setEnabled(self.model.is_data_valid())
        self.model.bad_rows_changed.connect(self.create_sip_button.setEnabled)
        
        grid_layout.addWidget(series_label, 0, 0, 1, 4)
        grid_layout.addWidget(self.default_sorting_button, 0, 4, 1, 1)
        grid_layout.addWidget(self.show_bad_rows_checkbox, 1, 0)
        grid_layout.addWidget(self.table_view, 2, 0, 1, 5)
        grid_layout.addWidget(save_button, 3, 0, 1, 2)
        grid_layout.addWidget(self.create_sip_button, 3, 2, 1, 3)

    def _bad_rows_clicked(self, state: QtCore.Qt.CheckState) -> None:
        if state == QtCore.Qt.CheckState.Checked.value:
            self.proxy_model.add_filter(CustomSortFilterModel.BAD_ROWS_FILTER)
        else:
            self.proxy_model.remove_filter(CustomSortFilterModel.BAD_ROWS_FILTER)

    def save_button_click(self) -> None:
        try:
            # Overwrites all the data, no recovery possible with current implementation
            with sql.connect(self.list_item.source_path) as conn:
                conn.execute("DROP TABLE data;")

                conn.execute(f"""
                    CREATE TABLE data (
                        _id INTEGER PRIMARY KEY,
                        
                        {',\n'.join(f'"{c}" TEXT' for c in self.list_item.grid.columns if c != '_id')}
                    );
                """)
                conn.executemany(f"INSERT INTO data VALUES ({'?,' * (len(self.list_item.grid.columns) - 1)}?)", self.list_item.grid.data)
        except Exception:
            WarningDialog(
                title="Ongekende fout",
                text="Ongekende fout is opgetreden tijdens het opslaan van de grid.",
            ).exec()
        
        if not self.saved:
            Dialog(
                title="Opgeslagen", text="De metadata is succesvol opgeslagen."
            ).exec()

        self.saved = True

    def create_sip_click(self, manual=True) -> None:
        def _col_index_to_xslx_col(col_index: int) -> str:
            # NOTE: this only supports up to ZZ for now
            first_letter_value = (col_index // 26) - 1

            if first_letter_value >= 26:
                raise ValueError("There are too many columns")
            if first_letter_value == -1:
                return chr(65 + col_index)

            first_letter = chr(65 + first_letter_value)
            second_letter = chr(65 + col_index % 26)
            
            return f"{first_letter}{second_letter}"

        self.save_button_click()

        self.state.configuration.create_locations()
        
        # Download the import_template
        import_template_loc = APIController.get_import_template(self.state.configuration, series_id=self.model.series._id)

        # Copy import_template to grid_storage
        temp_loc = os.path.join(self.state.configuration.grid_location, f"temp_{self.model.series._id}.xlsx")

        wb = load_workbook(import_template_loc)

        try:
            ws = wb["Details"]

            data = self.model.raw_data

            # Check for empty lines and drop them
            rows_to_drop = [
                r for r in data
                if all(
                    v == ""
                    for v in r[1:]
                )
            ]

            for r in rows_to_drop:
                data.remove(r)

            # NOTE: max lines alowed is 9999 (counting heading as well in the excel)
            if len(data) > 9999 - 1:
                WarningDialog(
                    title="Te veel lijnen",
                    text=f"Er zijn te veel lijnen in de tab '{self.model.series.name}'.\n\nMaximum: 9999\nGevonden: {len(data) + 1}"
                ).exec()
                return
            
            # NOTE: overwrite the headers since we might have added columns
            for col_index, col in self.model.columns.items():
                if col == "_id":
                    continue

                # Since we skipped a column, make sure to update the index
                col_index = col_index - 1

                ws[f"{_col_index_to_xslx_col(col_index)}1"] = col.strip()

            # NOTE: write the data
            for row_index, row in enumerate(data):
                # NOTE: since we need to start writing on row 2, update the index
                row_index = row_index + 2

                for col_index, value in enumerate(row):
                    if self.model.columns[col_index] == "_id":
                        continue

                    # Since we skipped a column, make sure to update the index
                    col_index = col_index - 1

                    ws[f"{_col_index_to_xslx_col(col_index)}{row_index}"] = value

            wb.save(temp_loc)
        finally:
            wb.close()

        # Create zipfile and md5
        name = self.list_item.name

        sip_location = os.path.join(self.state.configuration.sips_location, f"{self.model.series._id}-{name}-SIPC.zip")
        md5_location = os.path.join(self.state.configuration.sips_location, f"{self.model.series._id}-{name}-SIPC.xml")

        with zipfile.ZipFile(
            sip_location, "w", compression=zipfile.ZIP_DEFLATED
        ) as zfile:
            zfile.write(
                temp_loc,
                "Metadata.xlsx"
            )

        md5 = hashlib.md5(open(sip_location, "rb").read()).hexdigest()

        side_car_info = """<?xml version="1.0" encoding="UTF-8"?>
<mhs:Sidecar xmlns:mhs="https://zeticon.mediahaven.com/metadata/20.3/mhs/" version="20.3" xmlns:mh="https://zeticon.mediahaven.com/metadata/20.3/mh/">
<mhs:Technical>
        <mh:Md5>{md5}</mh:Md5>
</mhs:Technical>
</mhs:Sidecar>""".format(
            md5=md5
        )

        with open(md5_location, "w", encoding="utf-8") as f:
            f.write(side_car_info)

        os.remove(temp_loc)

        if manual:
            Dialog(
                title="SIPs aangemaakt",
                text="SIPS zijn aangemaakt voor de overdrachtslijst."
            ).exec()

            self.close()

    def closeEvent(self, event):
        if not self.saved:
            dialog = YesNoDialog(
                title="Opslaan",
                text="Gemaakte wijzigingen opslaan?"
            )
            dialog.exec()

            if dialog.result():
                self.save_button_click()

        event.accept()
