import re

import openpyxl
import pandas as pd

from src.utils.constants import (
    DATE_FORMAT,
    OVERDRACHTSLIJST_SHEET_NAME,
    UI_TEXT_ELEMENTS,
    OverdrachtslijstColumnName,
)


def _format_number(value, number_format: str) -> str:
    """Apply an Excel number_format to a numeric value to get the display text"""
    if number_format in (None, "General", "general"):
        # General format: Excel displays integers without decimals
        if isinstance(value, float) and value == int(value):
            return str(int(value))

        return str(value)

    if number_format == "0":
        return str(int(round(value)))

    # Percentage: "0%", "0.00%"  — must be checked before decimal places
    # because "0.00%" also contains a decimal group.
    if "%" in number_format:
        decimal_match = re.search(r"\.([0#]+)", number_format)
        decimal_places = len(decimal_match.group(1)) if decimal_match else 0
        return f"{value * 100:.{decimal_places}f}%"

    # Fixed decimal places: "0.00", "0.0", "#,##0.00", etc.
    decimal_match = re.search(r"\.([0#]+)", number_format)

    if decimal_match:
        decimal_places = len(decimal_match.group(1))
        return f"{value:.{decimal_places}f}"

    # Fallback: same as General
    if isinstance(value, float) and value == int(value):
        return str(int(value))

    return str(value)


def _format_cell(cell) -> str:
    """Convert an openpyxl cell to its display string."""
    value = cell.value

    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime(DATE_FORMAT)
    if isinstance(value, (int, float)):
        return _format_number(value, cell.number_format)

    return str(value)


def _warn_user(title: str, text: str) -> None:
    from PySide6.QtWidgets import QApplication

    app = QApplication.instance()

    if app and hasattr(app, "notify_user_signal"):
        app.notify_user_signal.emit(title, text)


def _deduplicate_headers(raw_headers: list[str]) -> list[str]:
    """Deduplicate column names by appending trailing spaces to repeats."""
    seen: dict[str, int] = {}
    headers = []

    for h in raw_headers:
        if h in seen:
            seen[h] += 1
            headers.append(h + " " * seen[h])
        else:
            seen[h] = 0
            headers.append(h)

    return headers


def _read_headers_from_anchor(row, start_col: int) -> list[str]:
    """Read header names from a row starting at start_col, stopping at the first empty cell.

    Returns deduplicated headers.
    """
    raw = []

    for cell in row[start_col:]:
        if cell.value is None:
            break
        raw.append(str(cell.value))

    return _deduplicate_headers(raw)


def _find_anchor_cell(rows, anchor: str) -> tuple[int, int] | None:
    """Find the row and column index of the anchor cell. Returns None if not found."""
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            if cell.value == anchor:
                return row_idx, col_idx

    return None


def _open_workbook(path: str) -> openpyxl.Workbook | None:
    """Open an Excel workbook, warning the user on failure. Returns None if it cannot be opened."""
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except Exception:
        _warn_user(
            UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["title"],
            UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["text"].format(path=path),
        )
        return None


class ExcelController:
    @staticmethod
    def read_excel(path: str) -> pd.DataFrame | None:
        """Reads an Excel file and returns its contents as a DataFrame.

        All values are read as display text, matching what Excel shows.
        Dates are formatted as YYYY-MM-DD.

        Returns None and warns the user if the file cannot be read.
        """
        wb = _open_workbook(path)

        if wb is None:
            return None

        try:
            ws = wb.active
            rows = list(ws.iter_rows())

            if not rows:
                _warn_user(
                    UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["title"],
                    UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["text"].format(path=path),
                )
                return None

            raw_headers = [str(cell.value) if cell.value is not None else "" for cell in rows[0]]

            # Strip trailing empty columns (phantom columns from Excel formatting)
            while raw_headers and raw_headers[-1] == "":
                raw_headers.pop()

            if not raw_headers or "" in raw_headers:
                _warn_user(
                    UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["title"],
                    UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["text"].format(path=path),
                )
                return None

            headers = _deduplicate_headers(raw_headers)
            num_columns = len(headers)

            data = [[_format_cell(cell) for cell in row[:num_columns]] for row in rows[1:]]

            # Strip trailing rows where every cell is empty
            while data and all(cell == "" for cell in data[-1]):
                data.pop()

            return pd.DataFrame(data, columns=headers).fillna("")
        except Exception:
            _warn_user(
                UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["title"],
                UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["text"].format(path=path),
            )
            return None
        finally:
            wb.close()

    @staticmethod
    def read_overdrachtslijst(path: str) -> pd.DataFrame | None:
        """Read an overdrachtslijst Excel file.

        Opens the sheet named 'Overdrachtslijst', finds the header row by
        scanning for 'Doosnr', reads columns to the right until an empty
        header, then reads data rows below until an entirely empty row.

        Returns None and warns the user on failure.
        """
        wb = _open_workbook(path)

        if wb is None:
            return None

        ui_errors = UI_TEXT_ELEMENTS["errors"]["migration"]
        anchor = OverdrachtslijstColumnName.DOOSNR

        try:
            if OVERDRACHTSLIJST_SHEET_NAME not in wb.sheetnames:
                _warn_user(
                    ui_errors["sheet_not_found"]["title"],
                    ui_errors["sheet_not_found"]["text"].format(
                        sheet_name=OVERDRACHTSLIJST_SHEET_NAME,
                        available_sheets=", ".join(wb.sheetnames),
                    ),
                )
                return None

            ws = wb[OVERDRACHTSLIJST_SHEET_NAME]
            rows = list(ws.iter_rows())

            if not rows:
                _warn_user(
                    ui_errors["sheet_empty"]["title"],
                    ui_errors["sheet_empty"]["text"].format(sheet_name=OVERDRACHTSLIJST_SHEET_NAME),
                )
                return None

            anchor_pos = _find_anchor_cell(rows, anchor)

            if anchor_pos is None:
                _warn_user(
                    ui_errors["anchor_not_found"]["title"],
                    ui_errors["anchor_not_found"]["text"].format(
                        anchor=anchor,
                        sheet_name=OVERDRACHTSLIJST_SHEET_NAME,
                    ),
                )
                return None

            anchor_row, anchor_col = anchor_pos
            headers = _read_headers_from_anchor(rows[anchor_row], anchor_col)

            if not headers:
                _warn_user(
                    ui_errors["no_columns"]["title"],
                    ui_errors["no_columns"]["text"].format(anchor=anchor),
                )
                return None

            num_columns = len(headers)
            data = []

            for row in rows[anchor_row + 1 :]:
                row_cells = row[anchor_col : anchor_col + num_columns]
                row_data = [_format_cell(cell) for cell in row_cells]

                if all(v == "" for v in row_data):
                    break

                data.append(row_data)

            return pd.DataFrame(data, columns=headers).fillna("")
        except Exception:
            _warn_user(
                UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["title"],
                UI_TEXT_ELEMENTS["errors"]["excel"]["read_error"]["text"].format(path=path),
            )
            return None
        finally:
            wb.close()
