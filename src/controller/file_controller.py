import os
import re
import hashlib
import shutil
import zipfile

import pandas as pd
from openpyxl import load_workbook

from src.controller.api_controller import APIController
from src.utils.base_object import BaseObject
from src.utils.constants import ColumnName, RowType, UI_TEXT_ELEMENTS
from src.utils.data_objects.digital.sip import SIP

UI_TEXT = UI_TEXT_ELEMENTS["errors"]["sip"]


SIDECAR_TEMPLATE = """<?xml version="1.0" encoding="UTF-8"?>
<mhs:Sidecar xmlns:mhs="https://zeticon.mediahaven.com/metadata/20.3/mhs/" version="20.3" xmlns:mh="https://zeticon.mediahaven.com/metadata/20.3/mh/">
     <mhs:Technical>
              <mh:Md5>{md5}</mh:Md5>
     </mhs:Technical>
</mhs:Sidecar>"""


class FileController(BaseObject):
    @staticmethod
    def _col_index_to_xlsx_col(col_index: int) -> str:
        result = ""
        index = col_index

        while index >= 0:
            result = chr(65 + index % 26) + result
            index = index // 26 - 1

        return result

    @staticmethod
    def _fill_import_template(df: pd.DataFrame, path: str) -> None:
        wb = load_workbook(path)
        ws = wb["Details"]

        for i, col in enumerate(df.columns):
            re_match = re.match(r"(.*)(\.\d+| +)$", col)

            if re_match:
                col = re_match.group(1)

            ws[f"{FileController._col_index_to_xlsx_col(i)}1"] = col

        for row_index, row_info in df.iterrows():
            for col_index, value in enumerate(row_info.values):
                ws[f"{FileController._col_index_to_xlsx_col(col_index)}{row_index+2}"] = value

        wb.save(path)
        wb.close()

    def _is_sip_folder_structure_valid(self, sip_folder_structure: dict, df: pd.DataFrame) -> bool:
        folder_paths = set()

        for row in sip_folder_structure.values():
            path_in_sip, path = row[ColumnName.PATH_IN_SIP.value], row["path"]

            if sum(df[ColumnName.PATH_IN_SIP.value] == path_in_sip) == 0:
                self.application.notify_user_signal.emit(
                    UI_TEXT["folder_structure_new_path_error"]["title"],
                    UI_TEXT["folder_structure_new_path_error"]["text"].format(path=path),
                )
                return False

            elif sum(df[ColumnName.PATH_IN_SIP.value] == path_in_sip) > 1:
                self.application.notify_user_signal.emit(
                    UI_TEXT["folder_structure_duplicate_path_error"]["title"],
                    UI_TEXT["folder_structure_duplicate_path_error"]["text"].format(path_in_sip=path_in_sip),
                )
                return False

            folder_paths.add(path_in_sip)

        for _, row in df.iterrows():
            if row[ColumnName.PATH_IN_SIP.value] not in folder_paths:
                self.application.notify_user_signal.emit(
                    UI_TEXT["folder_structure_missing_path_error"]["title"],
                    UI_TEXT["folder_structure_missing_path_error"]["text"].format(path_in_sip=row[ColumnName.PATH_IN_SIP.value]),
                )
                return False

        return True

    @staticmethod
    def _filter_df(df: pd.DataFrame, strip_name_extensions: bool = False) -> pd.DataFrame:
        filtered = df.loc[df[ColumnName.TYPE.value] != RowType.GEEN].copy()
        filtered.reset_index(drop=True, inplace=True)

        if strip_name_extensions and ColumnName.NAAM.value in filtered.columns:
            filtered[ColumnName.NAAM.value] = filtered[ColumnName.NAAM.value].apply(
                lambda v: v.rsplit(".", 1)[0] if v else v
            )

        return filtered

    def create_sip(self, sip: SIP, strip_name_extensions: bool = False) -> bool:
        configuration = self.application.configuration

        storage_location = configuration.sips_location
        import_template_location = os.path.join(
            configuration.import_templates_location,
            f"{sip.series._id}.xlsx"
        )

        os.makedirs(storage_location, exist_ok=True)

        if not os.path.exists(import_template_location):
            import_template_location = APIController.get_import_template(
                configuration=configuration,
                series_id=sip.series._id,
                environment=sip.environment,
            )

        sip_folder_structure = sip._get_folder_structure()
        filtered_folder_structure = {
            k: v for k, v in sip_folder_structure.items()
            if v[ColumnName.TYPE.value] != RowType.GEEN
        }

        df = FileController._filter_df(sip.grid_data.data_as_df, strip_name_extensions)

        if not self._is_sip_folder_structure_valid(
            sip_folder_structure=filtered_folder_structure,
            df=df
        ):
            return False

        temp_excel_location = os.path.join(
            configuration.import_templates_location,
            "temp.xlsx"
        )
        shutil.copy(
            src=import_template_location,
            dst=temp_excel_location
        )

        FileController._fill_import_template(df, temp_excel_location)

        sip_location = os.path.join(storage_location, sip.file_name)
        sidecar_location = os.path.join(storage_location, sip.sidecar_file_name)

        with zipfile.ZipFile(sip_location, "w", compression=zipfile.ZIP_DEFLATED) as zfile:
            zfile.write(temp_excel_location, "Metadata.xlsx")

            for location in filtered_folder_structure.values():
                zfile.write(location["path"], location[ColumnName.PATH_IN_SIP.value])

        with open(sip_location, "rb") as f:
            md5 = hashlib.md5(f.read()).hexdigest()

        with open(sidecar_location, "w", encoding="utf-8") as f:
            f.write(SIDECAR_TEMPLATE.format(md5=md5))

        try:
            os.remove(temp_excel_location)
        except PermissionError:
            import gc
            gc.collect()
            os.remove(temp_excel_location)

        return True
