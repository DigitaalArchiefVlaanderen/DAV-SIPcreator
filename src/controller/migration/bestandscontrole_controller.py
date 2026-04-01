import os
import re

import pandas as pd
from openpyxl import load_workbook

from src.utils.base_object import ApplicationMixin
from src.utils.constants import ColumnName

LIST_SHEET_NAME = "Lijst"
DOOS_SHEET_NAME = "Doostypes"

LIST_NAME_COLUMN = "Lijst benaming"
LIST_START_COLUMN = "Locatie dozen: begin\n(verdiep/blok*rij*rek*ligger)"
LIST_END_COLUMN = "Locatie dozen: eind\n(verdiep/blok*rij*rek*ligger)"
LIST_TYPE_COLUMN = "Doostype\n(1, 2, 3)\n(indien meerdere types: meest voorkomende)"

LIST_COLUMN_NAMES = [
    LIST_NAME_COLUMN,
    LIST_START_COLUMN,
    LIST_END_COLUMN,
    LIST_TYPE_COLUMN,
]

DOOS_NUMBER_COLUMN = "Doosnummer"
DOOS_TYPE_COLUMN = "Naam doostype migratie"

DOOS_COLUMN_NAMES = [
    DOOS_NUMBER_COLUMN,
    DOOS_TYPE_COLUMN,
]

VALUE_TO_COLUMN = {
    LIST_START_COLUMN: ColumnName.LEGACY_LOCATIE_ID.value,
    LIST_END_COLUMN: ColumnName.LEGACY_RANGE.value,
    DOOS_TYPE_COLUMN: ColumnName.VERPAKKINGSTYPE.value,
}


class BestandsControleController(ApplicationMixin):
    def __init__(self) -> None:
        super().__init__()

        self.controle_list_path: str = ""
        self.valid: bool = False
        self.list_df: pd.DataFrame | None = None

    def reset(self, path: str) -> None:
        if path == self.controle_list_path:
            return

        self.controle_list_path = path
        self.valid = False
        self.list_df = None

        self._load()

    def _notify(self, title: str, text: str) -> None:
        self.application.notify_user_signal.emit(title, text)

    def _check_valid(self) -> bool:
        if self.controle_list_path in ("", "."):
            return False

        if not os.path.exists(self.controle_list_path):
            self._notify(
                "Bestandscontrole niet gevonden",
                f"Bestandscontrole niet gevonden op locatie '{self.controle_list_path}'.",
            )
            return False

        wb = load_workbook(
            self.controle_list_path,
            read_only=True,
            data_only=True,
            keep_links=False,
            rich_text=False,
        )

        try:
            return self._validate_workbook(wb)
        finally:
            wb.close()

    def _validate_workbook(self, wb) -> bool:
        if LIST_SHEET_NAME not in wb.sheetnames:
            self._notify(
                "Tab niet gevonden in bestandscontrole",
                f"Tab '{LIST_SHEET_NAME}' niet gevonden in bestandscontrole.",
            )
            return False

        if DOOS_SHEET_NAME not in wb.sheetnames:
            self._notify(
                "Tab niet gevonden in bestandscontrole",
                f"Tab '{DOOS_SHEET_NAME}' niet gevonden in bestandscontrole.",
            )
            return False

        if not self._validate_sheet_headers(wb[LIST_SHEET_NAME], "Datum", LIST_COLUMN_NAMES, LIST_SHEET_NAME):
            return False

        if not self._validate_sheet_headers(wb[DOOS_SHEET_NAME], "Doosnummer", DOOS_COLUMN_NAMES, DOOS_SHEET_NAME):
            return False

        return True

    def _validate_sheet_headers(self, sheet, marker: str, required_columns: tuple, sheet_name: str) -> bool:
        data = sheet.values

        try:
            while marker not in (headers := next(data)):
                pass
        except StopIteration:
            self._notify(
                "Verwachte hoofdingen niet gevonden",
                f"De verwachte hoofdingen waren niet gevonden in de bestandscontrole in tab '{sheet_name}'.",
            )
            return False

        headers = [h for h in headers if h is not None]

        if any(h not in headers for h in required_columns):
            self._notify(
                "Verwachte hoofdingen niet gevonden",
                f"De verwachte hoofdingen waren niet gevonden in de bestandscontrole in tab '{sheet_name}'.",
            )
            return False

        return True

    def _load(self) -> None:
        if not self._check_valid():
            self.valid = False

            return

        self.valid = True

        wb = load_workbook(
            self.controle_list_path,
            read_only=True,
            data_only=True,
            keep_links=False,
            rich_text=False,
        )

        ws_list = wb[LIST_SHEET_NAME]
        list_data = list(ws_list.values)

        list_data_header_row = [i for i, r in enumerate(list_data) if all(c in r for c in LIST_COLUMN_NAMES)][0]
        list_data_columns = [i for i, c in enumerate(list_data[list_data_header_row]) if c in LIST_COLUMN_NAMES]
        list_data = [
            [c for i, c in enumerate(r) if i in list_data_columns] for r in list_data[list_data_header_row + 1 :]
        ]

        list_df = (
            pd.DataFrame(
                [r for r in list_data if not all(c is None for c in r)],
                columns=LIST_COLUMN_NAMES,
            )
            .fillna("")
            .astype(str)
            .convert_dtypes()
        )

        ws_doos = wb[DOOS_SHEET_NAME]
        doos_data = list(ws_doos.values)

        doos_data_header_row = [i for i, r in enumerate(doos_data) if all(c in r for c in DOOS_COLUMN_NAMES)][0]
        doos_data_columns = [i for i, c in enumerate(doos_data[doos_data_header_row]) if c in DOOS_COLUMN_NAMES]
        doos_data = [
            [c for i, c in enumerate(r) if i in doos_data_columns] for r in doos_data[doos_data_header_row + 1 :]
        ]

        doos_df = (
            pd.DataFrame(
                [r for r in doos_data if not all(c is None for c in r)],
                columns=DOOS_COLUMN_NAMES,
            )
            .fillna("")
            .astype(str)
            .convert_dtypes()
        )

        merged_df = list_df.merge(doos_df, left_on=LIST_TYPE_COLUMN, right_on=DOOS_NUMBER_COLUMN, how="left")
        merged_df[DOOS_TYPE_COLUMN] = merged_df[DOOS_TYPE_COLUMN].fillna(merged_df[LIST_TYPE_COLUMN])

        self.list_df = merged_df[[LIST_NAME_COLUMN, LIST_START_COLUMN, LIST_END_COLUMN, DOOS_TYPE_COLUMN]]

        wb.close()

    def get_values(self, overdrachtslijst_name: str) -> dict | None:
        self._load()

        if not self.valid or self.list_df is None:
            return None

        trimmed_name = re.split(r"(?i)_klant", overdrachtslijst_name)[0]

        output = self.list_df.loc[self.list_df[LIST_NAME_COLUMN] == trimmed_name]

        if len(output) == 0:
            self._notify(
                "Overdrachtslijst niet gevonden in bestandscontrole",
                f"Overdrachtslijst '{trimmed_name}' is niet gevonden in de kolom '{LIST_NAME_COLUMN}' van de bestandscontrole.",
            )

            return None

        if len(output) != 1:
            self._notify(
                "Overdrachtslijst te vaak gevonden in bestandscontrole",
                f"Overdrachtslijst '{trimmed_name}' is te vaak gevonden in de kolom '{LIST_NAME_COLUMN}' van de bestandscontrole.\n\nEr zijn {len(output)} rijen die over dezelfde overdrachtslijst gaan.",
            )

            return None

        for col in (LIST_START_COLUMN, LIST_END_COLUMN, DOOS_TYPE_COLUMN):
            if output.at[output.index[0], col] in (None, ""):
                self._notify(
                    "Lege waarde gevonden",
                    f"De kolom '{col}' bevat een lege waarde voor overdrachtslijst '{trimmed_name}'.",
                )

        return output.fillna("").to_dict(orient="records")[0]
