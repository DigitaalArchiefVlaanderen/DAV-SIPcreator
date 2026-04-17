"""Shared SIP creation utilities for all application types.

Provides the common operations for building SIP ZIP files:
- Filling an import template with grid data
- Creating a ZIP with Metadata.xlsx (and optional additional files)
- Generating the MD5 sidecar XML
"""

import hashlib
import os
import re
import zipfile

from openpyxl import load_workbook

SIDECAR_TEMPLATE = """<?xml version="1.0" encoding="UTF-8"?>
<mhs:Sidecar xmlns:mhs="https://zeticon.mediahaven.com/metadata/20.3/mhs/" version="20.3" xmlns:mh="https://zeticon.mediahaven.com/metadata/20.3/mh/">
     <mhs:Technical>
              <mh:Md5>{md5}</mh:Md5>
     </mhs:Technical>
</mhs:Sidecar>"""

COLUMN_NAME_CLEANUP_REGEX = re.compile(r"(.*)(\.\d+| +)$")


def fill_import_template(df, template_path: str, output_path: str) -> None:
    """Fill an import template Excel file with grid data.

    Opens the template, writes cleaned column names and data rows to the
    'Details' sheet, and saves to output_path.
    """
    wb = load_workbook(template_path)

    try:
        ws = wb["Details"]

        for col_index, col_name in enumerate(df.columns):
            clean_name = col_name.strip()
            match = COLUMN_NAME_CLEANUP_REGEX.match(clean_name)

            if match:
                clean_name = match.group(1)

            ws.cell(row=1, column=col_index + 1, value=clean_name)

        for row_index in range(len(df)):
            for col_index in range(len(df.columns)):
                ws.cell(row=row_index + 2, column=col_index + 1, value=str(df.iat[row_index, col_index]))

        wb.save(output_path)
    finally:
        wb.close()


def create_sip_zip(
    metadata_path: str,
    sip_location: str,
    sidecar_location: str,
    additional_files: dict[str, str] | None = None,
) -> None:
    """Create a SIP ZIP file with Metadata.xlsx and optional additional files.

    Args:
        metadata_path: Path to the filled Metadata.xlsx file.
        sip_location: Output path for the ZIP file.
        sidecar_location: Output path for the sidecar XML file.
        additional_files: Optional dict mapping {archive_name: disk_path} for
            extra files to include in the ZIP (used by digital SIPs for
            dossier files).
    """
    with zipfile.ZipFile(sip_location, "w", compression=zipfile.ZIP_DEFLATED) as zfile:
        zfile.write(metadata_path, "Metadata.xlsx")

        if additional_files:
            for archive_name, disk_path in additional_files.items():
                zfile.write(disk_path, archive_name)

    with open(sip_location, "rb") as f:
        md5 = hashlib.md5(f.read()).hexdigest()

    with open(sidecar_location, "w", encoding="utf-8") as f:
        f.write(SIDECAR_TEMPLATE.format(md5=md5))


def create_simple_sip(sip, configuration, df=None) -> bool:
    """Create a SIP ZIP for analog or migration (Metadata.xlsx only, no files).

    Uses sip.grid_data.data_as_df if df is not provided.
    Filters out empty rows (where all cells are empty string).
    Returns True on success.
    """
    from src.controller.api_controller import APIController

    if df is None:
        df = sip.grid_data.data_as_df

    non_empty_mask = df.apply(lambda row: any(str(v) != "" for v in row), axis=1)
    non_empty_df = df[non_empty_mask].reset_index(drop=True)

    configuration.create_locations()

    import_template_loc = APIController.get_import_template(
        configuration=configuration,
        environment=sip.environment,
        series_id=sip.series._id,
    )

    temp_loc = os.path.join(configuration.grid_location, f"temp_{sip.series._id}.xlsx")
    sip_location = os.path.join(configuration.sips_location, sip.file_name)
    sidecar_location = os.path.join(configuration.sips_location, sip.sidecar_file_name)

    fill_import_template(non_empty_df, import_template_loc, temp_loc)

    try:
        create_sip_zip(temp_loc, sip_location, sidecar_location)
    finally:
        os.remove(temp_loc)

    return True


def create_migration_series_sips(sip, configuration, series_data: list) -> bool:
    """Create SIP ZIPs for migration series.

    Args:
        sip: The MigrationSIP.
        configuration: Application configuration.
        series_data: List of (series_name, series_id, df) tuples.

    Returns True on success.
    """
    from src.controller.api_controller import APIController

    from src.utils.constants import BusinessRules

    configuration.create_locations()
    ol_name = sip.name[: BusinessRules.SIP_TITLE_MAX_LENGTH]

    for _, series_id, df in series_data:
        import_template_loc = APIController.get_import_template(
            configuration=configuration,
            environment=sip.environment,
            series_id=series_id,
        )

        temp_loc = os.path.join(configuration.grid_location, f"temp_{series_id}.xlsx")

        sip_file_name = f"{series_id}-{ol_name}-SIPC.zip"
        sidecar_file_name = f"{series_id}-{ol_name}-SIPC.xml"

        sip_location = os.path.join(configuration.sips_location, sip_file_name)
        sidecar_location = os.path.join(configuration.sips_location, sidecar_file_name)

        fill_import_template(df, import_template_loc, temp_loc)

        try:
            create_sip_zip(temp_loc, sip_location, sidecar_location)
        finally:
            os.remove(temp_loc)

    return True


def update_metadata_in_zip(metadata_path: str, sip_location: str, sidecar_location: str) -> None:
    """Replace only the Metadata.xlsx inside an existing ZIP and regenerate the sidecar.

    Used when the grid data changed but the additional files (digital) haven't.
    Falls back to full recreation if the ZIP doesn't exist.
    """
    if not os.path.exists(sip_location):
        create_sip_zip(metadata_path, sip_location, sidecar_location)
        return

    temp_zip = sip_location + ".tmp"

    with (
        zipfile.ZipFile(sip_location, "r") as zin,
        zipfile.ZipFile(temp_zip, "w", compression=zipfile.ZIP_DEFLATED) as zout,
    ):
        for item in zin.infolist():
            if item.filename == "Metadata.xlsx":
                continue
            zout.writestr(item, zin.read(item.filename))

        zout.write(metadata_path, "Metadata.xlsx")

    os.replace(temp_zip, sip_location)

    with open(sip_location, "rb") as f:
        md5 = hashlib.md5(f.read()).hexdigest()

    with open(sidecar_location, "w", encoding="utf-8") as f:
        f.write(SIDECAR_TEMPLATE.format(md5=md5))
