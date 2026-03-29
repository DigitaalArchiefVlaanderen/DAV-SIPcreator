import hashlib
import os
import re
import zipfile

from openpyxl import load_workbook
from PySide6 import QtWidgets, QtCore

from src.controller.api_controller import APIController
from src.controller.excel_controller import ExcelController
from src.utils.constants import ColumnName, UI_TEXT_ELEMENTS, BusinessRules
from src.utils.data_objects.analog.sip import AnalogSIP
from src.utils.data_objects.sip_status import SIPStatus
from src.utils.grid.table.analog_data_verification_table import (
    AnalogDataVerificationTable, NON_DUPLICATABLE_COLUMNS,
)
from src.utils.grid.table.common.grid_table_view import GridTableView
from src.utils.grid.table.common.proxy_model import SortFilterProxyModel, TableFilter
from src.utils.pyside_helper import set_widget_warning_style, clear_widget_warning_style
from src.utils.workers.worker import Worker

from src.widget.base_widget import BaseWidget

UI_TEXT = UI_TEXT_ELEMENTS["analog"]["grid"]
COMMON_GRID_TEXT = UI_TEXT_ELEMENTS["grid_checks"]["common"]

SIDECAR_TEMPLATE = """<?xml version="1.0" encoding="UTF-8"?>
<mhs:Sidecar xmlns:mhs="https://zeticon.mediahaven.com/metadata/20.3/mhs/" version="20.3" xmlns:mh="https://zeticon.mediahaven.com/metadata/20.3/mh/">
<mhs:Technical>
        <mh:Md5>{md5}</mh:Md5>
</mhs:Technical>
</mhs:Sidecar>"""


class AnalogGridView(BaseWidget):
    def __init__(self, sip: AnalogSIP) -> None:
        super().__init__()

        self.sip = sip
        self.has_unsaved_changes = False

        self.setup_ui()
        self.setup_signals()

    def setup_ui(self) -> None:
        self.grid_layout = QtWidgets.QGridLayout()
        self.setLayout(self.grid_layout)

        self.series_label = QtWidgets.QLabel()

        self.default_sorting_button = QtWidgets.QPushButton(text=UI_TEXT["default_sorting_button_text"])

        self.show_bad_rows_checkbox = QtWidgets.QCheckBox(text=UI_TEXT["bad_rows_checkbox_text"])

        self.column_dropdown = QtWidgets.QComboBox()
        self.column_dropdown.setEditable(True)
        self.column_dropdown.setInsertPolicy(QtWidgets.QComboBox.InsertPolicy.NoInsert)
        self.add_column_button = QtWidgets.QPushButton(text=UI_TEXT["add_column_button_text"])

        self.row_count_label = QtWidgets.QLabel()

        self.table_model = AnalogDataVerificationTable(sip=self.sip)
        self.proxy_model = SortFilterProxyModel()
        self.proxy_model.setSourceModel(self.table_model)

        self.table_view = GridTableView()
        self.table_view.setModel(self.proxy_model)

        self._populate_column_dropdown()
        self.column_dropdown.completer().setFilterMode(QtCore.Qt.MatchFlag.MatchContains)

        self.save_button = QtWidgets.QPushButton(text=UI_TEXT["save_button_text"])
        self.create_sip_button = QtWidgets.QPushButton(text=UI_TEXT["create_sip_button_text"])

        self._update_series_label()
        self._update_row_count_label(self.table_model.count_data_rows())

        if self.sip.series:
            self.table_model.validate_all()

        controls_layout = QtWidgets.QHBoxLayout()
        controls_layout.addWidget(self.show_bad_rows_checkbox)
        controls_layout.addStretch()
        controls_layout.addWidget(self.column_dropdown)
        controls_layout.addWidget(self.add_column_button)
        controls_layout.addWidget(self.row_count_label)

        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addWidget(self.save_button, 1)
        button_layout.addWidget(self.create_sip_button, 1)

        self.grid_layout.addWidget(self.series_label, 0, 0, 1, 4)
        self.grid_layout.addWidget(self.default_sorting_button, 0, 4, 1, 1)
        self.grid_layout.addLayout(controls_layout, 1, 0, 1, 5)
        self.grid_layout.addWidget(self.table_view, 2, 0, 1, 5)
        self.grid_layout.addLayout(button_layout, 3, 0, 1, 5)

    def setup_signals(self) -> None:
        self.default_sorting_button.clicked.connect(self.proxy_model.reset_sorting)
        self.show_bad_rows_checkbox.stateChanged.connect(self._bad_rows_clicked)
        self.add_column_button.clicked.connect(self._add_column_button_clicked)
        self.save_button.clicked.connect(self._save_button_clicked)
        self.create_sip_button.clicked.connect(self._create_sip_clicked)
        self.table_model.dataChanged.connect(self._data_changed)
        self.table_model.data_rows_changed_signal.connect(self._update_row_count_label)
        self.table_model.validation_started_signal.connect(self._update_create_sip_button)
        self.table_model.validation_finished_signal.connect(self._update_create_sip_button)
        self.sip.series_changed_signal.connect(self._on_series_changed)

    def _on_series_changed(self) -> None:
        self._update_series_label()

        if self.sip.series:
            self.table_model.validate_all()

    def _update_series_label(self) -> None:
        if self.sip.series:
            self.series_label.setText(self.sip.series.get_full_name())
            clear_widget_warning_style(self.series_label)
        else:
            fallback = self.sip.saved_series_name or self.sip.name
            self.series_label.setText(fallback)
            set_widget_warning_style(self.series_label, COMMON_GRID_TEXT["series_not_found_tooltip"])

        self._update_create_sip_button()

    def _populate_column_dropdown(self) -> None:
        seen = set()

        for col in self.table_model.raw_data.columns:
            base_col = col.rstrip()

            if base_col not in NON_DUPLICATABLE_COLUMNS and base_col not in seen:
                self.column_dropdown.addItem(base_col)
                seen.add(base_col)

    def _update_create_sip_button(self) -> None:
        self.create_sip_button.setEnabled(
            self.table_model.is_data_valid()
            and self.sip.series is not None
            and not self.table_model.is_validating
        )

    def _update_row_count_label(self, count: int) -> None:
        self.row_count_label.setText(UI_TEXT["row_count_label"].format(count=count))

        if count > BusinessRules.MAX_ROWS_PER_SERIES:
            self.row_count_label.setStyleSheet("QLabel {color: red;}")
        else:
            self.row_count_label.setStyleSheet("QLabel {color: black;}")

    def _data_changed(self) -> None:
        self.has_unsaved_changes = True
        self._update_create_sip_button()

    def _bad_rows_clicked(self, state: int) -> None:
        self.proxy_model.toggle_filter(TableFilter.BAD_ROWS)

    def _add_column_button_clicked(self) -> None:
        column = self.column_dropdown.currentText()

        if not column:
            return

        self.table_model.insert_column(column)

    def _save_button_clicked(self, silent: bool = False) -> None:
        self.application.analog_sip_db_controller.save_data(self.sip, self.table_model.raw_data)
        self.has_unsaved_changes = False

        if not silent:
            self.application.notify_user_signal.emit(
                UI_TEXT["save_success"]["title"],
                UI_TEXT["save_success"]["text"],
            )

    def _create_sip_clicked(self) -> None:
        self._save_button_clicked(silent=True)

        self.save_button.setEnabled(False)
        self.create_sip_button.setEnabled(False)

        self._create_sip_worker = Worker.start(
            self._background_create_sip,
            on_result=self._on_sip_created,
            on_error=lambda e: self.application.error_handler(e),
            on_finished=self._on_create_sip_finished,
        )

    def _background_create_sip(self) -> bool:
        non_empty_df = self.table_model.get_non_empty_df()

        if len(non_empty_df) > BusinessRules.MAX_ROWS_PER_SERIES:
            self.application.notify_user_signal.emit(
                UI_TEXT["too_many_rows_error"]["title"],
                UI_TEXT["too_many_rows_error"]["text"].format(
                    max_rows=BusinessRules.MAX_ROWS_PER_SERIES,
                    found_rows=len(non_empty_df)
                ),
            )

            return False

        self.application.configuration.create_locations()

        import_template_loc = APIController.get_import_template(
            configuration=self.application.configuration,
            environment=self.sip.environment,
            series_id=self.sip.series._id,
        )

        temp_loc = os.path.join(
            self.application.configuration.grid_location,
            f"temp_{self.sip.series._id}.xlsx"
        )

        wb = load_workbook(import_template_loc)

        try:
            ws = wb["Details"]

            for col_index, col_name in enumerate(non_empty_df.columns):
                clean_name = col_name.strip()
                matches = re.match(r"(.+)\.\d+", clean_name)

                if matches is not None:
                    clean_name = matches.group(1)

                ws.cell(row=1, column=col_index + 1, value=clean_name)

            for row_index in range(len(non_empty_df)):
                for col_index in range(len(non_empty_df.columns)):
                    ws.cell(
                        row=row_index + 2,
                        column=col_index + 1,
                        value=str(non_empty_df.iat[row_index, col_index])
                    )

            wb.save(temp_loc)
        finally:
            wb.close()

        sip_location = os.path.join(
            self.application.configuration.sips_location,
            self.sip.file_name
        )
        md5_location = os.path.join(
            self.application.configuration.sips_location,
            self.sip.sidecar_file_name
        )

        with zipfile.ZipFile(sip_location, "w", compression=zipfile.ZIP_DEFLATED) as zfile:
            zfile.write(temp_loc, "Metadata.xlsx")

        with open(sip_location, "rb") as f:
            md5 = hashlib.md5(f.read()).hexdigest()

        with open(md5_location, "w", encoding="utf-8") as f:
            f.write(SIDECAR_TEMPLATE.format(md5=md5))

        os.remove(temp_loc)

        return True

    def _on_sip_created(self, success: bool) -> None:
        if not success:
            return

        self.sip.set_status(SIPStatus.SIP_CREATED)

        self.application.notify_user_signal.emit(
            UI_TEXT["create_sip_success"]["title"],
            UI_TEXT["create_sip_success"]["text"],
        )

    def _on_create_sip_finished(self) -> None:
        self.save_button.setEnabled(True)
        self._update_create_sip_button()
