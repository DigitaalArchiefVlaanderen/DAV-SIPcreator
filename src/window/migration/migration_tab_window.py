import hashlib
import os
import re
import zipfile

import pandas as pd
from openpyxl import load_workbook
from PySide6 import QtCore, QtGui, QtWidgets

from src.controller.api_controller import APIController
from src.controller.excel_controller import ExcelController
from src.controller.file_controller import SIDECAR_TEMPLATE
from src.utils.constants import (
    ColumnName, UI_TEXT_ELEMENTS, OverdrachtslijstColumnName, RowType, BusinessRules,
    ANALOOG_DEFAULT_VALUE, MIGRATION_MAIN_ID_COLUMN, SERIES_NAME_COLUMN,
    DBColumnName,
)
from src.utils.data_objects.grid_data import GridData
from src.utils.data_objects.migration.sip import MigrationSIP
from src.utils.data_objects.sip_status import SIPStatus
from src.utils.pyside_helper import set_widget_warning_style, clear_widget_warning_style
from src.utils.workers.worker import Worker

from src.widget.central_widgets.migration.migration_grid_view import MigrationGridView
from src.widget.central_widgets.migration.migration_main_tab_view import MigrationMainTabView
from src.widget.dialog.yes_no_dialog import YesNoDialog

from src.window.base_window import Window


UI_TEXT = UI_TEXT_ELEMENTS["migration"]["tab_window"]

MAIN_TO_SERIES_COLUMN_MAPPING = {
    OverdrachtslijstColumnName.BESCHRIJVING.value: [ColumnName.NAAM.value, ColumnName.PATH_IN_SIP.value],
    OverdrachtslijstColumnName.BEGINDATUM.value: [ColumnName.OPENINGSDATUM.value],
    OverdrachtslijstColumnName.EINDDATUM.value: [ColumnName.SLUITINGSDATUM.value],
    OverdrachtslijstColumnName.DOOSNR.value: [ColumnName.ORIGINEEL_DOOSNUMMER.value],
}

FIXED_VALUE_COLUMNS = {
    ColumnName.ANALOOG.value: ANALOOG_DEFAULT_VALUE,
}

URI_SERIEREGISTER_COLUMN = DBColumnName.URI_SERIEREGISTER.value


class MigrationTabWindow(Window):
    def __init__(self, sip: MigrationSIP) -> None:
        super().__init__()

        self.sip = sip
        self.series_tabs: dict[str, MigrationGridView] = {}
        self._active_workers: list[tuple[Worker, QtCore.QThread]] = []

        self.setWindowTitle(self.sip.name)
        self.resize(1200, 800)

        self.setup_ui()
        self.setup_signals()
        self._load_existing_series_tabs()

    def setup_ui(self) -> None:
        self.tab_widget = QtWidgets.QTabWidget()
        self.setCentralWidget(self.tab_widget)

        self._ensure_series_name_column()

        self.sip.grid_data = self.sip.main_grid_data
        self.main_tab_view = MigrationMainTabView(sip=self.sip)
        self.tab_widget.addTab(self.main_tab_view, UI_TEXT["main_tab_title"])

    def setup_signals(self) -> None:
        self.main_tab_view.assign_to_series_signal.connect(self._assign_rows_to_series)
        self.main_tab_view.create_sip_signal.connect(self._create_all_sips)
        self.sip.name_changed_signal.connect(lambda: self.setWindowTitle(self.sip.name))
        self.application.application_environment_changed_signal.connect(self.close)

    def _ensure_series_name_column(self) -> None:
        df = self.sip.main_grid_data.data_as_df
        changed = False

        if SERIES_NAME_COLUMN not in df.columns:
            df.insert(0, SERIES_NAME_COLUMN, "")
            changed = True

        if URI_SERIEREGISTER_COLUMN not in df.columns:
            insert_pos = 1 if SERIES_NAME_COLUMN in df.columns else 0
            df.insert(insert_pos, URI_SERIEREGISTER_COLUMN, "")
            changed = True

        if changed:
            self.sip.main_grid_data.data_as_df = df

    def _load_existing_series_tabs(self) -> None:
        self._set_tab_loading(0, True)
        self._set_controls_busy(True)

        def background_load_tabs():
            tables = self.application.migration_sip_db_controller.read_tables(self.sip.db_name)
            results = []

            for table_name, uri_serieregister, edepot_id, status in tables:
                df = self.application.migration_sip_db_controller.read_series_data(self.sip.db_name, table_name)
                results.append((table_name, df))

            return results

        worker = Worker(function=background_load_tabs, is_generator=False)
        thread = QtCore.QThread()

        worker.moveToThread(thread)
        self._active_workers.append((worker, thread))

        thread.started.connect(worker.run)
        worker.result_ready_signal.connect(self._on_tabs_loaded)
        worker.error_encountered_signal.connect(
            lambda e: self.application.error_handler(e)
        )
        worker.finished_signal.connect(thread.quit)
        thread.finished.connect(thread.deleteLater)
        worker.finished_signal.connect(lambda: self._on_load_worker_finished(worker, thread))

        thread.start()

    def _on_tabs_loaded(self, results: list[tuple[str, pd.DataFrame]]) -> None:
        existing_table_names = set()

        for table_name, df in results:
            existing_table_names.add(table_name)

            grid_data = GridData()
            grid_data.data_as_df = df
            self.sip.series_grid_data[table_name] = grid_data

            grid_view = MigrationGridView(sip=self.sip, series_name=table_name, grid_data=grid_data)
            grid_view.table_model.validation_finished_signal.connect(self.update_global_create_sip_button)
            grid_view.create_sip_signal.connect(self._create_all_sips)
            self.series_tabs[table_name] = grid_view

            tab_index = self.tab_widget.addTab(grid_view, table_name)
            self._set_tab_loading(tab_index, True)

        self._create_missing_series_tabs(existing_table_names)

    def _on_load_worker_finished(self, worker: Worker, thread: QtCore.QThread) -> None:
        self._active_workers.remove((worker, thread))
        self._set_controls_busy(False)

        for i in range(self.tab_widget.count()):
            self._set_tab_loading(i, False)

    def _set_tab_loading(self, tab_index: int, loading: bool) -> None:
        tab_bar = self.tab_widget.tabBar()

        if loading:
            tab_bar.setTabTextColor(tab_index, QtGui.QColor("red"))
            current_text = self.tab_widget.tabText(tab_index)

            if not current_text.endswith(UI_TEXT["tab_loading_text"]):
                self.tab_widget.setTabText(tab_index, f"{current_text} ({UI_TEXT['tab_loading_text']})")
        else:
            tab_bar.setTabTextColor(tab_index, QtGui.QColor())
            current_text = self.tab_widget.tabText(tab_index)
            suffix = f" ({UI_TEXT['tab_loading_text']})"

            if current_text.endswith(suffix):
                self.tab_widget.setTabText(tab_index, current_text[:-len(suffix)])

    def _create_missing_series_tabs(self, existing_table_names: set[str]) -> None:
        main_df = self.sip.main_grid_data.data_as_df

        if SERIES_NAME_COLUMN not in main_df.columns or URI_SERIEREGISTER_COLUMN not in main_df.columns:
            return

        name_col = main_df.columns.get_loc(SERIES_NAME_COLUMN)
        uri_col = main_df.columns.get_loc(URI_SERIEREGISTER_COLUMN)

        unlinked_series: dict[str, dict] = {}

        for row in range(main_df.shape[0]):
            series_name = str(main_df.iat[row, name_col]).strip()
            uri = str(main_df.iat[row, uri_col]).strip()

            if not series_name or series_name == "nan":
                continue

            if series_name in existing_table_names:
                continue

            if series_name not in unlinked_series:
                unlinked_series[series_name] = {"rows": [], "uri": uri}

            unlinked_series[series_name]["rows"].append(row)

        for series_name, info in unlinked_series.items():
            uri = info["uri"]
            series_id = uri.rsplit("/", 1)[-1] if uri and uri != "nan" else ""

            template_columns = self._get_template_columns(series_id) if series_id else None
            series_df = self._map_main_to_series(main_df.iloc[info["rows"]].copy(), template_columns)

            uri_serieregister = uri if uri and uri != "nan" else ""

            grid_data = GridData()
            grid_data.data_as_df = series_df
            self.sip.series_grid_data[series_name] = grid_data

            self.application.migration_sip_db_controller.create_series_table(
                self.sip, uri_serieregister=uri_serieregister, table_name=series_name, df=series_df
            )

            grid_view = MigrationGridView(sip=self.sip, series_name=series_name, grid_data=grid_data)
            grid_view.table_model.validation_finished_signal.connect(self.update_global_create_sip_button)
            grid_view.create_sip_signal.connect(self._create_all_sips)
            self.series_tabs[series_name] = grid_view
            self.tab_widget.addTab(grid_view, series_name)

    def _get_template_columns(self, series_id: str) -> list[str] | None:
        storage_location = self.application.configuration.misc.save_location
        file_location = os.path.join(storage_location, "import_templates", f"{series_id}.xlsx")

        if not os.path.exists(file_location):
            try:
                file_location = APIController.get_import_template(
                    configuration=self.application.configuration,
                    environment=self.sip.environment,
                    series_id=series_id,
                )
            except Exception:
                return None

        template_df = ExcelController.read_excel(file_location)

        if template_df is None:
            return None

        seen: dict[str, int] = {}
        columns = []

        for col in template_df.columns:
            base_col = re.sub(r"\.\d+$", "", col)

            if base_col in seen:
                seen[base_col] += 1
                col = base_col + " " * seen[base_col]
            else:
                seen[base_col] = 0
                col = base_col

            columns.append(col)

        return columns

    def _set_controls_busy(self, busy: bool) -> None:
        self.main_tab_view.assign_button.setEnabled(not busy)
        self.main_tab_view.save_button.setEnabled(not busy)
        self.main_tab_view.create_sip_button.setEnabled(not busy)

        for grid_view in self.series_tabs.values():
            grid_view.save_button.setEnabled(not busy)
            grid_view.create_sip_button.setEnabled(not busy)

        if busy:
            set_widget_warning_style(
                self.main_tab_view.assign_button,
                tooltip=UI_TEXT["assign_loading_tooltip"],
            )
        else:
            clear_widget_warning_style(self.main_tab_view.assign_button)
            self.main_tab_view._update_create_sip_button()

            for grid_view in self.series_tabs.values():
                grid_view._update_create_sip_button()

    def _assign_rows_to_series(self, source_rows: list[int], series_id: str, series_name: str) -> None:
        table_name = series_name
        uri_serieregister = f"{self.sip.environment.get_serie_register_uri()}/{series_id}"
        main_df = self.sip.main_grid_data.data_as_df

        existing_count = 0

        if table_name in self.series_tabs:
            existing_count = len(self.series_tabs[table_name].table_model.raw_data)

        total_count = existing_count + len(source_rows)

        if total_count > BusinessRules.MAX_ROWS_PER_SERIES:
            self.application.notify_user_signal.emit(
                UI_TEXT["row_limit_error"]["title"],
                UI_TEXT["row_limit_error"]["text"].format(
                    max_rows=BusinessRules.MAX_ROWS_PER_SERIES,
                    total_rows=total_count,
                ),
            )

            return

        selected_data = main_df.iloc[source_rows].copy()

        self._remove_rows_from_old_series(main_df, source_rows, table_name)

        main_df.iloc[source_rows, main_df.columns.get_loc(SERIES_NAME_COLUMN)] = series_name
        main_df.iloc[source_rows, main_df.columns.get_loc(URI_SERIEREGISTER_COLUMN)] = uri_serieregister
        self.sip.main_grid_data.data_as_df = main_df

        self.sip.grid_data = self.sip.main_grid_data

        self.main_tab_view.table_model.beginResetModel()
        self.main_tab_view.table_model.raw_data = main_df
        self.main_tab_view.table_model.endResetModel()

        self.main_tab_view._validate_series()

        self._set_controls_busy(True)

        def background_assign():
            self.application.migration_sip_db_controller.save_main_data(self.sip, main_df)
            template_columns = self._get_template_columns(series_id)
            series_df = self._map_main_to_series(selected_data, template_columns)

            return series_df, template_columns

        worker = Worker(function=background_assign, is_generator=False)
        thread = QtCore.QThread()

        worker.moveToThread(thread)
        self._active_workers.append((worker, thread))

        thread.started.connect(worker.run)
        worker.result_ready_signal.connect(
            lambda result: self._on_assign_complete(result, table_name, uri_serieregister, series_name)
        )
        worker.error_encountered_signal.connect(
            lambda e: self._on_assign_error(e, worker, thread)
        )
        worker.finished_signal.connect(thread.quit)
        thread.finished.connect(thread.deleteLater)
        worker.finished_signal.connect(lambda: self._on_assign_worker_finished(worker, thread))

        thread.start()

    def _on_assign_complete(self, result: tuple, table_name: str, uri_serieregister: str, series_name: str) -> None:
        series_df, template_columns = result

        if table_name in self.series_tabs:
            existing_grid_data = self.sip.series_grid_data[table_name]
            existing_df = existing_grid_data.data_as_df

            combined_df = pd.concat([existing_df, series_df], ignore_index=True)
            existing_grid_data.data_as_df = combined_df

            grid_view = self.series_tabs[table_name]
            grid_view.table_model.beginResetModel()
            grid_view.table_model.raw_data = combined_df
            grid_view.table_model.endResetModel()

            grid_view.table_model.re_mark_disabled_columns()
            grid_view.table_model.validate_all()

            self.application.migration_sip_db_controller.save_series_data(
                self.sip, table_name, combined_df
            )
        else:
            grid_data = GridData()
            grid_data.data_as_df = series_df
            self.sip.series_grid_data[table_name] = grid_data

            self.application.migration_sip_db_controller.create_series_table(
                self.sip, uri_serieregister=uri_serieregister, table_name=table_name, df=series_df
            )

            grid_view = MigrationGridView(sip=self.sip, series_name=series_name, grid_data=grid_data)
            grid_view.table_model.validation_finished_signal.connect(self.update_global_create_sip_button)
            grid_view.create_sip_signal.connect(self._create_all_sips)
            self.series_tabs[table_name] = grid_view
            self.tab_widget.addTab(grid_view, series_name)

    def _on_assign_error(self, error: Exception, worker: Worker, thread: QtCore.QThread) -> None:
        self.application.notify_user_signal.emit(
            UI_TEXT["assign_error"]["title"],
            UI_TEXT["assign_error"]["text"],
        )

    def _on_assign_worker_finished(self, worker: Worker, thread: QtCore.QThread) -> None:
        self._active_workers.remove((worker, thread))
        self._set_controls_busy(False)

    def _remove_rows_from_old_series(self, main_df: pd.DataFrame, source_rows: list[int], new_table_name: str) -> None:
        if SERIES_NAME_COLUMN not in main_df.columns:
            return

        name_col = main_df.columns.get_loc(SERIES_NAME_COLUMN)
        main_ids_by_old_series: dict[str, list[str]] = {}

        for row_idx in source_rows:
            old_series = str(main_df.iat[row_idx, name_col]).strip()

            if not old_series or old_series == "nan" or old_series == new_table_name:
                continue

            if old_series not in main_ids_by_old_series:
                main_ids_by_old_series[old_series] = []

            main_ids_by_old_series[old_series].append(str(row_idx))

        for old_series_name, main_ids in main_ids_by_old_series.items():
            if old_series_name not in self.series_tabs:
                continue

            grid_view = self.series_tabs[old_series_name]
            old_df = grid_view.table_model.raw_data
            main_id_set = set(main_ids)

            remaining_df = old_df[~old_df[MIGRATION_MAIN_ID_COLUMN].astype(str).isin(main_id_set)].reset_index(drop=True)

            if remaining_df.empty:
                tab_index = self.tab_widget.indexOf(grid_view)
                self.tab_widget.removeTab(tab_index)
                grid_view.deleteLater()

                del self.series_tabs[old_series_name]
                del self.sip.series_grid_data[old_series_name]

                self.application.migration_sip_db_controller.delete_series_table(self.sip, old_series_name)
            else:
                grid_view.grid_data.data_as_df = remaining_df

                grid_view.table_model.beginResetModel()
                grid_view.table_model.raw_data = remaining_df
                grid_view.table_model.endResetModel()

                self.application.migration_sip_db_controller.save_series_data(
                    self.sip, old_series_name, remaining_df
                )

    def _map_main_to_series(self, selected_data: pd.DataFrame, template_columns: list[str] | None = None) -> pd.DataFrame:
        mapped_data: dict[str, list] = {}
        row_count = len(selected_data)

        mapped_data[MIGRATION_MAIN_ID_COLUMN] = selected_data.index.astype(str).tolist()

        for main_col, series_cols in MAIN_TO_SERIES_COLUMN_MAPPING.items():
            if main_col in selected_data.columns:
                values = selected_data[main_col].values.tolist()
            else:
                values = [""] * row_count

            for series_col in series_cols:
                mapped_data[series_col] = list(values)

        for col_name, fixed_value in FIXED_VALUE_COLUMNS.items():
            mapped_data[col_name] = [fixed_value] * row_count

        self._format_origineel_doosnummer(mapped_data)
        self._derive_type_and_dossier_ref(mapped_data, row_count)

        ordered_columns = [MIGRATION_MAIN_ID_COLUMN]

        if template_columns:
            for col in template_columns:
                if col not in ordered_columns:
                    ordered_columns.append(col)

        series_df = pd.DataFrame()

        for col in ordered_columns:
            if col in mapped_data:
                series_df[col] = mapped_data[col]
            else:
                series_df[col] = ""

        return series_df.fillna("").reset_index(drop=True)

    def _format_origineel_doosnummer(self, mapped_data: dict[str, list]) -> None:
        col = ColumnName.ORIGINEEL_DOOSNUMMER.value

        if col not in mapped_data:
            return

        overdrachtslijst_name = self.sip.overdrachtslijst_name

        for i, value in enumerate(mapped_data[col]):
            raw = str(value).strip()

            if not raw or raw == "nan":
                mapped_data[col][i] = ""
                continue

            if re.match(r"^\d+$", raw):
                raw = raw.zfill(4)

            mapped_data[col][i] = f"{raw}/{overdrachtslijst_name}"

    @staticmethod
    def _derive_type_and_dossier_ref(mapped_data: dict[str, list], row_count: int) -> None:
        path_col = ColumnName.PATH_IN_SIP.value

        if path_col not in mapped_data:
            return

        types = []
        refs = []

        for value in mapped_data[path_col]:
            value = str(value).strip()

            if not value or value == "nan":
                types.append("")
                refs.append("")
            elif "/" in value:
                types.append(RowType.STUK)
                refs.append(value.split("/", 1)[0])
            else:
                types.append(RowType.DOSSIER)
                refs.append(value)

        mapped_data[ColumnName.TYPE.value] = types
        mapped_data[ColumnName.DOSSIER_REF.value] = refs

    def update_global_create_sip_button(self) -> None:
        all_valid = (
            self.series_tabs
            and all(
                not gv.table_model.has_bad_rows
                and gv.series is not None
                and not gv.table_model.is_validating
                for gv in self.series_tabs.values()
            )
            and not self.main_tab_view.table_model.has_bad_rows
        )

        self.main_tab_view.create_sip_button.setEnabled(all_valid)

        for grid_view in self.series_tabs.values():
            grid_view.create_sip_button.setEnabled(all_valid)

    def _create_all_sips(self) -> None:
        for grid_view in self.series_tabs.values():
            grid_view._save_button_clicked(silent=True)

        self.main_tab_view._save_button_clicked()

        series_data: list[tuple[str, str, pd.DataFrame]] = []

        for series_name, grid_view in self.series_tabs.items():
            if grid_view.series is None:
                continue

            series_id = grid_view.series._id
            df = grid_view.table_model.raw_data.copy()

            if MIGRATION_MAIN_ID_COLUMN in df.columns:
                df = df.drop(columns=[MIGRATION_MAIN_ID_COLUMN])

            if len(df) > BusinessRules.MAX_ROWS_PER_SERIES:
                self.application.notify_user_signal.emit(
                    UI_TEXT["row_limit_error"]["title"],
                    UI_TEXT["row_limit_error"]["text"].format(
                        max_rows=BusinessRules.MAX_ROWS_PER_SERIES,
                        total_rows=len(df),
                    ),
                )

                return

            series_data.append((series_name, series_id, df))

        if not series_data:
            return

        self._set_controls_busy(True)

        def background_create_sips():
            self.application.configuration.create_locations()
            configuration = self.application.configuration
            ol_name = self.sip.overdrachtslijst_name[:185]

            for series_name, series_id, df in series_data:
                import_template_loc = APIController.get_import_template(
                    configuration=configuration,
                    environment=self.sip.environment,
                    series_id=series_id,
                )

                temp_loc = os.path.join(
                    configuration.grid_location,
                    f"temp_{series_id}.xlsx"
                )

                wb = load_workbook(import_template_loc)

                try:
                    ws = wb["Details"]

                    for col_index, col_name in enumerate(df.columns):
                        clean_name = col_name.strip()
                        matches = re.match(r"(.+)[\s.]\d+$", clean_name)

                        if matches is not None:
                            clean_name = matches.group(1)

                        ws.cell(row=1, column=col_index + 1, value=clean_name)

                    for row_index in range(len(df)):
                        for col_index in range(len(df.columns)):
                            ws.cell(
                                row=row_index + 2,
                                column=col_index + 1,
                                value=str(df.iat[row_index, col_index])
                            )

                    wb.save(temp_loc)
                finally:
                    wb.close()

                sip_file_name = f"{series_id}-{ol_name}-SIPC.zip"
                sidecar_file_name = f"{series_id}-{ol_name}-SIPC.xml"

                sip_location = os.path.join(configuration.sips_location, sip_file_name)
                sidecar_location = os.path.join(configuration.sips_location, sidecar_file_name)

                with zipfile.ZipFile(sip_location, "w", compression=zipfile.ZIP_DEFLATED) as zfile:
                    zfile.write(temp_loc, "Metadata.xlsx")

                with open(sip_location, "rb") as f:
                    md5 = hashlib.md5(f.read()).hexdigest()

                with open(sidecar_location, "w", encoding="utf-8") as f:
                    f.write(SIDECAR_TEMPLATE.format(md5=md5))

                os.remove(temp_loc)

            return True

        worker = Worker(function=background_create_sips, is_generator=False)
        thread = QtCore.QThread()

        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.result_ready_signal.connect(self._on_all_sips_created)
        worker.error_encountered_signal.connect(
            lambda e: self.application.error_handler(e)
        )
        worker.finished_signal.connect(thread.quit)
        thread.finished.connect(thread.deleteLater)
        worker.finished_signal.connect(lambda: self._set_controls_busy(False))

        self._active_workers.append((worker, thread))
        worker.finished_signal.connect(lambda: self._on_create_worker_finished(worker, thread))

        thread.start()

    def _on_all_sips_created(self, success: bool) -> None:
        if not success:
            return

        for series_name in self.series_tabs:
            self.sip.series_statuses[series_name] = SIPStatus.SIP_CREATED

            self.application.migration_sip_db_controller.update_series_status(
                self.sip, series_name, SIPStatus.SIP_CREATED
            )

        self.sip.set_status(SIPStatus.SIP_CREATED)

        self.application.notify_user_signal.emit(
            UI_TEXT["create_all_sips_success"]["title"],
            UI_TEXT["create_all_sips_success"]["text"],
        )

    def _on_create_worker_finished(self, worker: Worker, thread: QtCore.QThread) -> None:
        if (worker, thread) in self._active_workers:
            self._active_workers.remove((worker, thread))

    def closeEvent(self, event: QtGui.QCloseEvent) -> None:
        has_unsaved = any(
            grid_view.has_unsaved_changes
            for grid_view in self.series_tabs.values()
        )

        if has_unsaved:
            dialog = YesNoDialog(
                title=UI_TEXT["unsaved_changes_dialog"]["title"],
                text=UI_TEXT["unsaved_changes_dialog"]["text"],
            )
            dialog.exec()

            if dialog.result():
                for grid_view in self.series_tabs.values():
                    if grid_view.has_unsaved_changes:
                        grid_view._save_button_clicked(silent=True)

                self.application.notify_user_signal.emit(
                    UI_TEXT["save_all_success"]["title"],
                    UI_TEXT["save_all_success"]["text"],
                )

        super().closeEvent(event)
